from datetime import timedelta
from io import BytesIO
from typing import Any, Dict, List

from django.contrib.auth.models import User
from django.db.models import Avg, Count, F, Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from producer.models import (
    AuditLog,
    Customer,
    LedgerEntry,
    MarketplaceProduct,
    Order,
    Payment,
    Producer,
    Product,
    Sale,
    StockHistory,
)


class BusinessDataExporter:
    def __init__(self, user: User):
        self.user = user
        self.user_profile = user.user_profile
        self.shop_id = self.user_profile.shop_id if self.user_profile else None
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.export_date = timezone.now()
        self.row_count = {}

    @staticmethod
    def get_header_style():
        return {
            "font": Font(bold=True, color="FFFFFF", size=11),
            "fill": PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        }

    @staticmethod
    def get_subheader_style():
        return {
            "font": Font(bold=True, color="FFFFFF", size=10),
            "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        }

    @staticmethod
    def get_data_style(alternate: bool = False):
        fill_color = "E7E6E6" if alternate else "FFFFFF"
        return {
            "font": Font(size=10),
            "fill": PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        }

    @staticmethod
    def get_success_style():
        return {
            "font": Font(bold=True, color="FFFFFF", size=10),
            "fill": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        }

    @staticmethod
    def get_warning_style():
        return {
            "font": Font(bold=True, color="FFFFFF", size=10),
            "fill": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
            "border": Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            ),
        }

    @staticmethod
    def safe_cell_value(value: Any) -> Any:
        if value is None:
            return "N/A"
        # Handle UUID objects
        if hasattr(value, "hex"):
            return str(value)
        return value

    @staticmethod
    def apply_style(cell, style: Dict[str, Any]):
        if "font" in style:
            cell.font = style["font"]
        if "fill" in style:
            cell.fill = style["fill"]
        if "alignment" in style:
            cell.alignment = style["alignment"]
        if "border" in style:
            cell.border = style["border"]
        if "number_format" in style:
            cell.number_format = style["number_format"]

    # ==================== DATA GATHERING ====================

    def get_business_summary(self) -> Dict[str, Any]:
        producers = Producer.objects.filter(user=self.user)
        products = Product.objects.filter(user=self.user)
        orders = Order.objects.filter(user=self.user)
        sales = Sale.objects.filter(user=self.user)
        customers = Customer.objects.filter(user=self.user)
        marketplace_products = MarketplaceProduct.objects.filter(product__user=self.user)

        total_sales_amount = sales.aggregate(total=Sum("sale_price"))["total"] or 0
        total_orders = orders.aggregate(total=Sum("quantity"))["total"] or 0
        total_sales_quantity = sales.aggregate(total=Sum("quantity"))["total"] or 0

        return {
            "user_name": f"{self.user.first_name} {self.user.last_name}".strip() or self.user.username,
            "username": self.user.username,
            "email": self.user.email,
            "business_name": self.user_profile.registered_business_name if self.user_profile else "N/A",
            "business_type": self.user_profile.get_business_type_display() if self.user_profile else "N/A",
            "shop_id": self.shop_id,
            "b2b_verified": self.user_profile.b2b_verified if self.user_profile else False,
            "total_producers": producers.count(),
            "total_products": products.count(),
            "total_marketplace_products": marketplace_products.count(),
            "total_orders": orders.count(),
            "total_customers": customers.count(),
            "total_order_quantity": total_orders,
            "total_sales_quantity": total_sales_quantity,
            "total_sales_amount": float(total_sales_amount),
            "average_order_value": float(total_sales_amount / orders.count()) if orders.count() > 0 else 0,
            "export_date": self.export_date.strftime("%Y-%m-%d %H:%M:%S"),
        }

    def get_sales_metrics(self) -> Dict[str, Any]:
        """Get detailed sales metrics."""
        sales = Sale.objects.filter(user=self.user)
        last_7_days = timezone.now() - timedelta(days=7)
        last_30_days = timezone.now() - timedelta(days=30)
        last_90_days = timezone.now() - timedelta(days=90)

        metrics = {
            "total_sales": sales.count(),
            "total_revenue": float(sales.aggregate(total=Sum("sale_price"))["total"] or 0),
            "average_sale_price": float(sales.aggregate(avg=Avg("sale_price"))["avg"] or 0),
            "total_units_sold": sales.aggregate(total=Sum("quantity"))["total"] or 0,
            "average_units_per_sale": sales.aggregate(avg=Avg("quantity"))["avg"] or 0,
            "last_7_days_sales": sales.filter(sale_date__gte=last_7_days).count(),
            "last_7_days_revenue": float(
                sales.filter(sale_date__gte=last_7_days).aggregate(total=Sum("sale_price"))["total"] or 0
            ),
            "last_30_days_sales": sales.filter(sale_date__gte=last_30_days).count(),
            "last_30_days_revenue": float(
                sales.filter(sale_date__gte=last_30_days).aggregate(total=Sum("sale_price"))["total"] or 0
            ),
            "last_90_days_sales": sales.filter(sale_date__gte=last_90_days).count(),
            "last_90_days_revenue": float(
                sales.filter(sale_date__gte=last_90_days).aggregate(total=Sum("sale_price"))["total"] or 0
            ),
            "pending_payments": sales.filter(payment_status="pending").count(),
            "completed_payments": sales.filter(payment_status="paid").count(),
        }

        return metrics

    def get_financial_summary(self) -> Dict[str, Any]:
        ledger_entries = LedgerEntry.objects.filter(user=self.user)

        sales_revenue = ledger_entries.filter(account_type="SR").aggregate(total=Sum("amount"))["total"] or 0
        cogs = ledger_entries.filter(account_type="COGS").aggregate(total=Sum("amount"))["total"] or 0
        vat_payable = ledger_entries.filter(account_type="VAT_P").aggregate(total=Sum("amount"))["total"] or 0
        tds_payable = ledger_entries.filter(account_type="TDS").aggregate(total=Sum("amount"))["total"] or 0
        accounts_payable = ledger_entries.filter(account_type="AP").aggregate(total=Sum("amount"))["total"] or 0
        accounts_receivable = ledger_entries.filter(account_type="AR").aggregate(total=Sum("amount"))["total"] or 0

        gross_profit = float(sales_revenue) - float(cogs)
        net_profit = gross_profit - float(vat_payable) - float(tds_payable)
        gross_margin = (gross_profit / float(sales_revenue) * 100) if sales_revenue > 0 else 0

        return {
            "total_revenue": float(sales_revenue),
            "total_cogs": float(cogs),
            "gross_profit": gross_profit,
            "gross_margin_percent": gross_margin,
            "net_profit": net_profit,
            "vat_payable": float(vat_payable),
            "tds_payable": float(tds_payable),
            "accounts_payable": float(accounts_payable),
            "accounts_receivable": float(accounts_receivable),
        }

    def get_inventory_metrics(self) -> Dict[str, Any]:
        products = Product.objects.filter(user=self.user)
        stock_histories = StockHistory.objects.filter(user=self.user)

        total_stock = products.aggregate(total=Sum("stock"))["total"] or 0
        total_stock_value = sum(p.stock * float(p.price) for p in products if p.stock and p.price)

        low_stock = products.filter(stock__lt=F("reorder_level")).count()
        out_of_stock = products.filter(stock=0).count()
        overstock = products.filter(stock__gt=F("reorder_level") * 3).count()

        # Calculate turnover
        total_stock_out = stock_histories.aggregate(total=Sum("quantity_out"))["total"] or 0

        return {
            "total_stock_units": total_stock,
            "total_inventory_value": float(total_stock_value),
            "low_stock_items": low_stock,
            "out_of_stock_items": out_of_stock,
            "overstock_items": overstock,
            "total_stock_movements": stock_histories.count(),
            "total_units_moved": total_stock_out,
        }

    def get_customer_metrics(self) -> Dict[str, Any]:
        customers = Customer.objects.filter(user=self.user)
        orders = Order.objects.filter(user=self.user)

        total_credit_extended = customers.aggregate(total=Sum("credit_limit"))["total"] or 0
        total_outstanding = customers.aggregate(total=Sum("current_balance"))["total"] or 0

        return {
            "total_customers": customers.count(),
            "retailers": customers.filter(customer_type="Retailer").count(),
            "wholesalers": customers.filter(customer_type="Wholesaler").count(),
            "distributors": customers.filter(customer_type="Distributor").count(),
            "total_credit_extended": float(total_credit_extended),
            "total_outstanding": float(total_outstanding),
            "average_customer_balance": float(total_outstanding / customers.count()) if customers.count() > 0 else 0,
        }

    def get_product_performance(self) -> List[Dict[str, Any]]:
        sales_data = (
            Sale.objects.filter(user=self.user)
            .values("order__product")
            .annotate(
                total_quantity=Sum("quantity"),
                total_revenue=Sum(F("quantity") * F("sale_price")),
                sale_count=Count("id"),
            )
        )

        result = []
        for data in sales_data:
            product_id = data["order__product"]
            if product_id:
                product = Product.objects.get(id=product_id)
                result.append(
                    {
                        "product_id": product_id,
                        "product_name": product.name,
                        "sku": product.sku,
                        "total_sold": data["total_quantity"] or 0,
                        "total_revenue": float(data["total_revenue"] or 0),
                        "sale_count": data["sale_count"],
                        "current_stock": product.stock,
                        "price": float(product.price),
                    }
                )

        return sorted(result, key=lambda x: x["total_revenue"], reverse=True)

    def create_executive_summary_sheet(self):
        ws = self.workbook.create_sheet("Executive Summary", 0)
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25

        # Title
        title_cell = ws["A1"]
        title_cell.value = "BUSINESS INTELLIGENCE DASHBOARD"
        self.apply_style(title_cell, {**self.get_header_style(), "font": Font(bold=True, size=14, color="FFFFFF")})
        ws.merge_cells("A1:B1")

        # Business Info
        row = 3
        summary = self.get_business_summary()

        info_items = [
            ("Business Owner", summary["user_name"]),
            ("Business Name", summary["business_name"]),
            ("Business Type", summary["business_type"]),
            ("Shop ID", summary["shop_id"]),
            ("B2B Verified", "✓ Yes" if summary["b2b_verified"] else "✗ No"),
            ("Export Date", summary["export_date"]),
        ]

        for label, value in info_items:
            ws[f"A{row}"].value = self.safe_cell_value(label)
            ws[f"B{row}"].value = self.safe_cell_value(value)
            self.apply_style(ws[f"A{row}"], {**self.get_subheader_style(), "alignment": Alignment(horizontal="left")})
            style = self.get_success_style() if "✓" in str(value) else self.get_data_style()
            self.apply_style(ws[f"B{row}"], style)
            row += 1

        # KEY PERFORMANCE INDICATORS
        row += 2
        kpi_title = ws[f"A{row}"]
        kpi_title.value = "KEY PERFORMANCE INDICATORS"
        self.apply_style(kpi_title, self.get_header_style())
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        metrics = self.get_sales_metrics()
        financial = self.get_financial_summary()
        inventory = self.get_inventory_metrics()
        customer_metrics = self.get_customer_metrics()

        kpi_items = [
            ("SALES & REVENUE", ""),
            ("Total Sales Count", f"{metrics['total_sales']}"),
            ("Total Revenue", f"Rs. {metrics['total_revenue']:,.2f}"),
            ("Avg Sale Value", f"Rs. {metrics['average_sale_price']:,.2f}"),
            ("Units Sold", f"{metrics['total_units_sold']}"),
            ("Revenue (Last 30 Days)", f"Rs. {metrics['last_30_days_revenue']:,.2f}"),
            ("", ""),
            ("FINANCIAL HEALTH", ""),
            ("Gross Profit", f"Rs. {financial['gross_profit']:,.2f}"),
            ("Gross Margin %", f"{financial['gross_margin_percent']:.2f}%"),
            ("Net Profit", f"Rs. {financial['net_profit']:,.2f}"),
            ("Accounts Receivable", f"Rs. {financial['accounts_receivable']:,.2f}"),
            ("Accounts Payable", f"Rs. {financial['accounts_payable']:,.2f}"),
            ("", ""),
            ("INVENTORY MANAGEMENT", ""),
            ("Total Stock Units", f"{inventory['total_stock_units']}"),
            ("Inventory Value", f"Rs. {inventory['total_inventory_value']:,.2f}"),
            ("Low Stock Items", f"{inventory['low_stock_items']}"),
            ("Out of Stock Items", f"{inventory['out_of_stock_items']}"),
            ("Overstock Items", f"{inventory['overstock_items']}"),
            ("", ""),
            ("CUSTOMER METRICS", ""),
            ("Total Customers", f"{customer_metrics['total_customers']}"),
            ("Total Credit Extended", f"Rs. {customer_metrics['total_credit_extended']:,.2f}"),
            ("Outstanding Balance", f"Rs. {customer_metrics['total_outstanding']:,.2f}"),
        ]

        for idx, (label, value) in enumerate(kpi_items):
            ws[f"A{row}"].value = self.safe_cell_value(label)
            ws[f"B{row}"].value = self.safe_cell_value(value)

            if label and not value:  # Category headers
                self.apply_style(ws[f"A{row}"], {**self.get_subheader_style(), "alignment": Alignment(horizontal="left")})
                ws.merge_cells(f"A{row}:B{row}")
            elif label and value:  # Data rows
                self.apply_style(
                    ws[f"A{row}"], {**self.get_data_style(alternate=idx % 2 == 0), "font": Font(bold=False, size=10)}
                )
                self.apply_style(ws[f"B{row}"], self.get_data_style(alternate=idx % 2 == 0))

            row += 1

    def create_financial_analysis_sheet(self):
        ws = self.workbook.create_sheet("Financial Analysis")

        # Headers
        headers = ["Account Type", "Amount (Rs.)", "Debit/Credit", "Transaction Date", "Reference"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = self.safe_cell_value(header)
            self.apply_style(cell, self.get_header_style())
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        ledger_entries = LedgerEntry.objects.filter(user=self.user).order_by("-date")

        for row_num, entry in enumerate(ledger_entries, 2):
            data = [
                entry.get_account_type_display(),
                float(entry.amount),
                "Debit" if entry.debit else "Credit",
                entry.date.strftime("%Y-%m-%d"),
                str(entry.reference_id) if entry.reference_id else "N/A",
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = self.safe_cell_value(value)
                self.apply_style(cell, self.get_data_style(alternate=row_num % 2 == 0))
                if col_num == 2:
                    cell.number_format = "#,##0.00"

        # Summary section
        row = len(ledger_entries) + 3
        summary = self.get_financial_summary()

        summary_data = [
            ("Total Revenue", summary["total_revenue"]),
            ("Cost of Goods Sold", summary["total_cogs"]),
            ("Gross Profit", summary["gross_profit"]),
            ("Gross Margin %", f"{summary['gross_margin_percent']:.2f}%"),
            ("VAT Payable", summary["vat_payable"]),
            ("TDS Payable", summary["tds_payable"]),
            ("Net Profit", summary["net_profit"]),
        ]

        ws[f"A{row}"].value = "FINANCIAL SUMMARY"
        self.apply_style(ws[f"A{row}"], self.get_header_style())
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        for label, value in summary_data:
            ws[f"A{row}"].value = label
            ws[f"B{row}"].value = value
            self.apply_style(ws[f"A{row}"], self.get_subheader_style())
            self.apply_style(ws[f"B{row}"], self.get_data_style())
            row += 1

    def create_inventory_analytics_sheet(self):
        ws = self.workbook.create_sheet("Inventory Analytics")

        # Product inventory status
        headers = [
            "Product Name",
            "SKU",
            "Current Stock",
            "Reorder Level",
            "Status",
            "Stock Value (Rs.)",
            "Days of Stock",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = self.safe_cell_value(header)
            self.apply_style(cell, self.get_header_style())
            ws.column_dimensions[get_column_letter(col_num)].width = 18

        try:
            products = Product.objects.filter(user=self.user).order_by("-stock")
        except:
            products = Product.objects.none()

        for row_num, product in enumerate(products, 2):
            status = "In Stock"
            if product.stock == 0:
                status = "Out of Stock"
            elif product.stock < product.reorder_level:
                status = "Low Stock"
            elif product.stock > product.reorder_level * 3:
                status = "Overstock"

            stock_value = product.stock * float(product.price) if product.price else 0

            data = [
                product.name,
                product.sku,
                product.stock,
                product.reorder_level,
                status,
                float(stock_value),
                "N/A",
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = self.safe_cell_value(value)

                if col_num == 5:  # Status column
                    if status == "Out of Stock":
                        self.apply_style(cell, self.get_warning_style())
                    elif status == "Low Stock":
                        self.apply_style(cell, self.get_warning_style())
                    else:
                        self.apply_style(cell, self.get_success_style())
                else:
                    self.apply_style(cell, self.get_data_style(alternate=row_num % 2 == 0))

        # Stock history summary
        row = len(products) + 3
        stock_metrics = self.get_inventory_metrics()

        ws[f"A{row}"].value = "INVENTORY SUMMARY"
        self.apply_style(ws[f"A{row}"], self.get_header_style())
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        summary_items = [
            ("Total Stock Units", stock_metrics["total_stock_units"]),
            ("Inventory Value", f"Rs. {stock_metrics['total_inventory_value']:,.2f}"),
            ("Low Stock Items", stock_metrics["low_stock_items"]),
            ("Out of Stock Items", stock_metrics["out_of_stock_items"]),
            ("Stock Movements", stock_metrics["total_stock_movements"]),
        ]

        for label, value in summary_items:
            ws[f"A{row}"].value = label
            ws[f"B{row}"].value = value
            self.apply_style(ws[f"A{row}"], self.get_subheader_style())
            self.apply_style(ws[f"B{row}"], self.get_data_style())
            row += 1

    def create_sales_performance_sheet(self):
        ws = self.workbook.create_sheet("Sales Performance")

        products_perf = self.get_product_performance()

        headers = [
            "Product Name",
            "SKU",
            "Units Sold",
            "Total Revenue (Rs.)",
            "Avg Price (Rs.)",
            "Sale Count",
            "Current Stock",
            "Performance Score",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = self.safe_cell_value(header)
            self.apply_style(cell, self.get_header_style())
            ws.column_dimensions[get_column_letter(col_num)].width = 18

        total_revenue = sum(p["total_revenue"] for p in products_perf)

        for row_num, product in enumerate(products_perf, 2):
            revenue_pct = (product["total_revenue"] / total_revenue * 100) if total_revenue > 0 else 0
            avg_price = product["total_revenue"] / product["total_sold"] if product["total_sold"] > 0 else 0

            performance_score = (
                "★★★★★" if revenue_pct > 20 else "★★★★" if revenue_pct > 10 else "★★★" if revenue_pct > 5 else "★★"
            )

            data = [
                product["product_name"],
                product["sku"],
                product["total_sold"],
                product["total_revenue"],
                avg_price,
                product["sale_count"],
                product["current_stock"],
                performance_score,
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = self.safe_cell_value(value)

                if col_num in [4, 5]:  # Currency columns
                    cell.number_format = "#,##0.00"

                self.apply_style(cell, self.get_data_style(alternate=row_num % 2 == 0))

    def create_customer_analysis_sheet(self):
        ws = self.workbook.create_sheet("Customer Analysis")

        headers = [
            "Customer Name",
            "Type",
            "Contact",
            "Email",
            "Credit Limit (Rs.)",
            "Outstanding (Rs.)",
            "Credit Usage %",
            "Status",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = self.safe_cell_value(header)
            self.apply_style(cell, self.get_header_style())
            ws.column_dimensions[get_column_letter(col_num)].width = 18

        customers = Customer.objects.filter(user=self.user).order_by("-current_balance")

        for row_num, customer in enumerate(customers, 2):
            usage_pct = (customer.current_balance / customer.credit_limit * 100) if customer.credit_limit > 0 else 0
            status = "Healthy" if usage_pct < 50 else "Warning" if usage_pct < 80 else "Critical"

            data = [
                customer.name,
                customer.customer_type,
                customer.contact,
                customer.email,
                float(customer.credit_limit),
                float(customer.current_balance),
                f"{usage_pct:.1f}%",
                status,
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = self.safe_cell_value(value)

                if col_num == 8:  # Status
                    if status == "Healthy":
                        self.apply_style(cell, self.get_success_style())
                    elif status == "Warning":
                        self.apply_style(cell, self.get_warning_style())
                    else:
                        self.apply_style(cell, self.get_warning_style())
                else:
                    self.apply_style(cell, self.get_data_style(alternate=row_num % 2 == 0))

    def create_orders_and_sales_sheet(self):
        ws = self.workbook.create_sheet("Orders & Sales")

        headers = [
            "Order #",
            "Customer",
            "Product",
            "Qty",
            "Order Value (Rs.)",
            "Status",
            "Payment Status",
            "Order Date",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = self.safe_cell_value(header)
            self.apply_style(cell, self.get_header_style())
            ws.column_dimensions[get_column_letter(col_num)].width = 16

        orders = Order.objects.filter(user=self.user).select_related("customer", "product").order_by("-order_date")

        for row_num, order in enumerate(orders, 2):
            data = [
                order.order_number,
                order.customer.name if order.customer else "N/A",
                order.product.name if order.product else "N/A",
                order.quantity,
                float(order.total_price) if order.total_price else 0,
                order.status,
                order.payment_status if hasattr(order, "payment_status") else "N/A",
                order.order_date.strftime("%Y-%m-%d"),
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = self.safe_cell_value(value)

                if col_num == 6:  # Order status
                    if "delivered" in str(value).lower():
                        self.apply_style(cell, self.get_success_style())
                    elif "pending" in str(value).lower():
                        self.apply_style(cell, self.get_warning_style())
                    else:
                        self.apply_style(cell, self.get_data_style(alternate=row_num % 2 == 0))
                else:
                    self.apply_style(cell, self.get_data_style(alternate=row_num % 2 == 0))

    def create_audit_trail_sheet(self):
        ws = self.workbook.create_sheet("Audit Trail")

        headers = ["Transaction Type", "Reference ID", "Amount (Rs.)", "Date", "Entity ID"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = self.safe_cell_value(header)
            self.apply_style(cell, self.get_header_style())
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        audit_logs = AuditLog.objects.filter(user=self.user).order_by("-date")[:500]

        for row_num, log in enumerate(audit_logs, 2):
            data = [
                log.get_transaction_type_display(),
                str(log.reference_id) if log.reference_id else "N/A",
                float(log.amount) if log.amount else 0,
                log.date.strftime("%Y-%m-%d"),
                str(log.entity_id) if log.entity_id else "N/A",
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = self.safe_cell_value(value)
                self.apply_style(cell, self.get_data_style(alternate=row_num % 2 == 0))

    def create_summary_metrics_sheet(self):
        ws = self.workbook.create_sheet("Key Metrics", 1)

        metrics = self.get_sales_metrics()
        financial = self.get_financial_summary()
        inventory = self.get_inventory_metrics()
        customer = self.get_customer_metrics()
        summary = self.get_business_summary()

        row = 1
        data_sections = [
            (
                "SALES METRICS",
                [
                    ("Total Sales", metrics["total_sales"]),
                    ("Total Revenue", f"Rs. {metrics['total_revenue']:,.2f}"),
                    ("7-Day Revenue", f"Rs. {metrics['last_7_days_revenue']:,.2f}"),
                    ("30-Day Revenue", f"Rs. {metrics['last_30_days_revenue']:,.2f}"),
                    ("90-Day Revenue", f"Rs. {metrics['last_90_days_revenue']:,.2f}"),
                    ("Avg Sale Value", f"Rs. {metrics['average_sale_price']:,.2f}"),
                    ("Pending Payments", metrics["pending_payments"]),
                    ("Completed Payments", metrics["completed_payments"]),
                ],
            ),
            (
                "INVENTORY STATUS",
                [
                    ("Total Units", inventory["total_stock_units"]),
                    ("Stock Value", f"Rs. {inventory['total_inventory_value']:,.2f}"),
                    ("Low Stock Items", inventory["low_stock_items"]),
                    ("Out of Stock", inventory["out_of_stock_items"]),
                    ("Overstock Items", inventory["overstock_items"]),
                    ("Stock Movements", inventory["total_stock_movements"]),
                ],
            ),
            (
                "FINANCIAL HEALTH",
                [
                    ("Revenue", f"Rs. {financial['total_revenue']:,.2f}"),
                    ("Cost of Goods", f"Rs. {financial['total_cogs']:,.2f}"),
                    ("Gross Profit", f"Rs. {financial['gross_profit']:,.2f}"),
                    ("Gross Margin %", f"{financial['gross_margin_percent']:.2f}%"),
                    ("Net Profit", f"Rs. {financial['net_profit']:,.2f}"),
                    ("Receivables", f"Rs. {financial['accounts_receivable']:,.2f}"),
                    ("Payables", f"Rs. {financial['accounts_payable']:,.2f}"),
                ],
            ),
            (
                "BUSINESS OVERVIEW",
                [
                    ("Total Producers", summary["total_producers"]),
                    ("Total Products", summary["total_products"]),
                    ("Total Orders", summary["total_orders"]),
                    ("Total Customers", customer["total_customers"]),
                    ("Marketplace Products", summary["total_marketplace_products"]),
                    ("Credit Extended", f"Rs. {customer['total_credit_extended']:,.2f}"),
                    ("Outstanding Credit", f"Rs. {customer['total_outstanding']:,.2f}"),
                ],
            ),
        ]

        for section_title, section_data in data_sections:
            # Section header
            header_cell = ws[f"A{row}"]
            header_cell.value = section_title
            self.apply_style(header_cell, self.get_header_style())
            ws.merge_cells(f"A{row}:B{row}")
            row += 1

            # Section data
            for label, value in section_data:
                ws[f"A{row}"].value = self.safe_cell_value(label)
                ws[f"B{row}"].value = self.safe_cell_value(value)
                self.apply_style(ws[f"A{row}"], {**self.get_data_style(), "font": Font(bold=True)})
                self.apply_style(ws[f"B{row}"], self.get_data_style())
                row += 1

            row += 1  # Space between sections

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25

    def generate_export(self) -> BytesIO:
        self.create_executive_summary_sheet()
        self.create_summary_metrics_sheet()
        self.create_financial_analysis_sheet()
        self.create_inventory_analytics_sheet()
        self.create_sales_performance_sheet()
        self.create_customer_analysis_sheet()
        self.create_orders_and_sales_sheet()
        self.create_audit_trail_sheet()

        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)

        return output
