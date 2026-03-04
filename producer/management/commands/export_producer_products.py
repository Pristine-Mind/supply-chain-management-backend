import csv
import json
from datetime import datetime

from django.core.management.base import BaseCommand
from django.db.models import Prefetch

from producer.models import Producer, Product, ProductImage

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class Command(BaseCommand):
    help = "Export producer data with their products including cost price, brand, price, discounted price and images"

    def add_arguments(self, parser):
        parser.add_argument(
            "--producer-ids",
            type=str,
            default="76,73,70,58,42",
            help="Comma-separated producer IDs to export (default: 76,73,70,58,42)",
        )

        parser.add_argument(
            "--format",
            type=str,
            choices=["csv", "json", "console", "excel"],
            default="excel",
            help="Export format: csv, json, excel, or console (default: excel)",
        )

        parser.add_argument(
            "--output-file",
            type=str,
            help="Output file path (if not provided, uses timestamp-based filename)",
        )

        parser.add_argument(
            "--discount-percentage",
            type=float,
            default=13.0,
            help="Discount percentage to calculate discounted price (default: 12)",
        )

    def handle(self, *args, **options):
        # Parse producer IDs
        producer_ids = [int(id.strip()) for id in options["producer_ids"].split(",")]
        discount_percentage = options["discount_percentage"]
        format_choice = options["format"]

        # Check if Excel format is requested but library not available
        if format_choice == "excel" and not EXCEL_AVAILABLE:
            self.stdout.write(
                self.style.ERROR("Excel format requires openpyxl library. Install it with: pip install openpyxl")
            )
            return

        self.stdout.write(f"Exporting data for producer IDs: {producer_ids}")
        self.stdout.write(f"Discount percentage: {discount_percentage}%")

        # Query producers with their products and images
        producers = (
            Producer.objects.filter(id__in=producer_ids)
            .prefetch_related(
                Prefetch(
                    "product_set",
                    queryset=Product.objects.select_related("brand").prefetch_related("images"),
                    to_attr="products",
                )
            )
            .order_by("id")
        )

        if not producers:
            self.stdout.write(self.style.WARNING("No producers found with the specified IDs"))
            return

        # Prepare export data
        if format_choice == "excel":
            self._export_to_excel(producers, discount_percentage, options.get("output_file"))
        else:
            # For other formats, prepare flat data structure
            export_data = self._prepare_flat_data(producers, discount_percentage)

            if format_choice == "console":
                self._export_to_console(export_data)
            elif format_choice == "json":
                self._export_to_json(export_data, options.get("output_file"))
            elif format_choice == "csv":
                self._export_to_csv(export_data, options.get("output_file"))

        self.stdout.write(self.style.SUCCESS(f"Successfully exported data for {len(producers)} producers"))

    def _prepare_flat_data(self, producers, discount_percentage):
        """Prepare flat data structure for CSV/JSON/Console export"""
        export_data = []

        for producer in producers:
            producer_data = {
                "producer_id": producer.id,
                "producer_name": producer.name,
                "producer_email": producer.email,
                "producer_contact": producer.contact,
                "producer_address": producer.address,
                "producer_registration_number": producer.registration_number,
                "producer_created_at": producer.created_at.isoformat(),
            }

            if not producer.products:
                # Include producer even if no products
                product_data = {
                    "product_id": None,
                    "product_name": None,
                    "brand_name": None,
                    "cost_price": None,
                    "price": None,
                    "discounted_price": None,
                    "stock": None,
                    "sku": None,
                    "is_active": None,
                    "product_created_at": None,
                    "category": None,
                    "size": None,
                    "color": None,
                    "images": [],
                }
                export_data.append({**producer_data, **product_data})
            else:
                # Add each product for the producer
                for product in producer.products:
                    # Calculate discounted price (discount_percentage% of cost_price)
                    discounted_price = None
                    if product.cost_price:
                        discounted_price = round(product.cost_price * (discount_percentage / 100), 2)

                    # Get product images
                    product_images = []
                    for image in product.images.all():
                        product_images.append(
                            {
                                "image_url": image.image.url if image.image else None,
                                "alt_text": image.alt_text,
                                "created_at": image.created_at.isoformat(),
                            }
                        )

                    product_data = {
                        "product_id": product.id,
                        "product_name": product.name,
                        "brand_name": product.get_brand_name(),
                        "cost_price": product.cost_price,
                        "price": product.price,
                        "discounted_price": discounted_price,
                        "stock": product.stock,
                        "sku": product.sku,
                        "is_active": product.is_active,
                        "product_created_at": product.created_at.isoformat(),
                        "category": product.get_category_hierarchy(),
                        "size": product.get_size_display() if product.size else None,
                        "color": product.get_color_display() if product.color else None,
                        "images": product_images,
                    }

                    export_data.append({**producer_data, **product_data})

        return export_data

    def _export_to_excel(self, producers, discount_percentage, output_file=None):
        """Export data to Excel workbook with separate sheets for each producer"""
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"producer_products_export_{timestamp}.xlsx"

        # Create workbook
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        producer_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        producer_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Create summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_headers = [
            "Producer ID",
            "Producer Name",
            "Email",
            "Contact",
            "Registration Number",
            "Total Products",
            "Total Stock Value",
        ]

        for col, header in enumerate(summary_headers, 1):
            cell = summary_ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        summary_row = 2

        for producer in producers:
            # Calculate totals for summary
            total_products = len(producer.products) if producer.products else 0
            total_stock_value = (
                sum((product.price * product.stock) for product in producer.products if product.price and product.stock)
                if producer.products
                else 0
            )

            # Add to summary
            summary_data = [
                producer.id,
                producer.name,
                producer.email,
                producer.contact,
                producer.registration_number,
                total_products,
                f"${total_stock_value:.2f}",
            ]

            for col, value in enumerate(summary_data, 1):
                summary_ws.cell(row=summary_row, column=col, value=value)

            summary_row += 1

            # Create individual sheet for each producer
            sheet_name = f"Producer_{producer.id}_{producer.name[:15]}"  # Limit sheet name length
            # Replace invalid characters for sheet names and ensure <= 31 chars
            sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (" ", "-", "_")).rstrip()[:31]

            ws = wb.create_sheet(sheet_name)

            # Producer Information Section
            ws.cell(row=1, column=1, value="PRODUCER INFORMATION").fill = producer_fill
            ws.cell(row=1, column=1).font = producer_font
            ws.merge_cells("A1:H1")

            producer_info = [
                ("Producer ID:", producer.id),
                ("Name:", producer.name),
                ("Email:", producer.email),
                ("Contact:", producer.contact),
                ("Address:", producer.address),
                ("Registration Number:", producer.registration_number),
                ("Created At:", producer.created_at.strftime("%Y-%m-%d %H:%M:%S")),
                ("Service Radius (km):", producer.service_radius_km),
            ]

            row = 2
            for label, value in producer_info:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                row += 1

            # Products Section
            row += 1
            ws.cell(row=row, column=1, value="PRODUCTS").fill = producer_fill
            ws.cell(row=row, column=1).font = producer_font
            ws.merge_cells(f"A{row}:O{row}")

            row += 1

            # Product headers
            product_headers = [
                "Product ID",
                "Product Name",
                "Brand",
                "SKU",
                "Category",
                "Cost Price",
                "Selling Price",
                f"Discounted Price ({discount_percentage}%)",
                "Stock",
                "Reorder Level",
                "Size",
                "Color",
                "Status",
                "Created At",
                "Image Count",
            ]

            for col, header in enumerate(product_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            row += 1

            if producer.products:
                for product in producer.products:
                    # Calculate discounted price
                    discounted_price = None
                    if product.cost_price:
                        discounted_price = round(product.cost_price * (discount_percentage / 100), 2)

                    # Count images
                    image_count = product.images.count()

                    product_data = [
                        product.id,
                        product.name,
                        product.get_brand_name(),
                        product.sku,
                        product.get_category_hierarchy(),
                        f"${product.cost_price:.2f}" if product.cost_price else None,
                        f"${product.price:.2f}" if product.price else None,
                        f"${discounted_price:.2f}" if discounted_price else None,
                        product.stock,
                        product.reorder_level,
                        product.get_size_display() if product.size else None,
                        product.get_color_display() if product.color else None,
                        "Active" if product.is_active else "Inactive",
                        product.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                        image_count,
                    ]

                    for col, value in enumerate(product_data, 1):
                        ws.cell(row=row, column=col, value=value)

                    row += 1

                # Add product images section if any products have images
                products_with_images = [p for p in producer.products if p.images.exists()]
                if products_with_images:
                    row += 2
                    ws.cell(row=row, column=1, value="PRODUCT IMAGES").fill = producer_fill
                    ws.cell(row=row, column=1).font = producer_font
                    ws.merge_cells(f"A{row}:D{row}")

                    row += 1

                    image_headers = ["Product ID", "Product Name", "Image URL", "Alt Text"]
                    for col, header in enumerate(image_headers, 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment

                    row += 1

                    for product in products_with_images:
                        for image in product.images.all():
                            image_data = [
                                product.id,
                                product.name,
                                image.image.url if image.image else "No URL",
                                image.alt_text or "No alt text",
                            ]

                            for col, value in enumerate(image_data, 1):
                                ws.cell(row=row, column=col, value=value)

                            row += 1
            else:
                ws.cell(row=row, column=1, value="No products found for this producer")

            # Auto-adjust column widths
            for column_cells in ws.columns:
                # Skip merged cells when calculating column width
                valid_cells = [cell for cell in column_cells if hasattr(cell, "column_letter") and cell.value]
                if valid_cells:
                    length = max(len(str(cell.value)) for cell in valid_cells)
                    ws.column_dimensions[valid_cells[0].column_letter].width = min(length + 2, 50)

        # Auto-adjust summary sheet column widths
        for column_cells in summary_ws.columns:
            # Skip merged cells when calculating column width
            valid_cells = [cell for cell in column_cells if hasattr(cell, "column_letter") and cell.value]
            if valid_cells:
                length = max(len(str(cell.value)) for cell in valid_cells)
                summary_ws.column_dimensions[valid_cells[0].column_letter].width = min(length + 2, 30)

        # Save workbook
        wb.save(output_file)
        self.stdout.write(f"Excel workbook exported to: {output_file}")
        self.stdout.write(f"Created {len(producers)} producer sheets plus summary sheet")

    def _export_to_console(self, data):
        """Export data to console"""
        self.stdout.write("\n" + "=" * 100)
        self.stdout.write("PRODUCER AND PRODUCT EXPORT DATA")
        self.stdout.write("=" * 100 + "\n")

        for item in data:
            self.stdout.write(f"Producer ID: {item['producer_id']}")
            self.stdout.write(f"Producer Name: {item['producer_name']}")
            self.stdout.write(f"Producer Email: {item['producer_email']}")
            self.stdout.write(f"Producer Contact: {item['producer_contact']}")

            if item["product_id"]:
                self.stdout.write(f"  Product ID: {item['product_id']}")
                self.stdout.write(f"  Product Name: {item['product_name']}")
                self.stdout.write(f"  Brand Name: {item['brand_name']}")
                self.stdout.write(f"  Cost Price: ${item['cost_price']}")
                self.stdout.write(f"  Price: ${item['price']}")
                self.stdout.write(f"  Discounted Price: ${item['discounted_price']}")
                self.stdout.write(f"  Stock: {item['stock']}")
                self.stdout.write(f"  Category: {item['category']}")
                self.stdout.write(f"  Images: {len(item['images'])} image(s)")
            else:
                self.stdout.write("  No products found for this producer")

            self.stdout.write("-" * 80)

    def _export_to_json(self, data, output_file=None):
        """Export data to JSON file"""
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"producer_products_export_{timestamp}.json"

        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        self.stdout.write(f"Data exported to: {output_file}")

    def _export_to_csv(self, data, output_file=None):
        """Export data to CSV file"""
        if not output_file:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"producer_products_export_{timestamp}.csv"

        if not data:
            self.stdout.write(self.style.WARNING("No data to export"))
            return

        # Flatten images data for CSV
        flattened_data = []
        for item in data:
            row = item.copy()
            # Convert images list to string representation
            if row["images"]:
                image_urls = [img["image_url"] for img in row["images"] if img["image_url"]]
                row["image_urls"] = "; ".join(image_urls)
                row["image_count"] = len(row["images"])
            else:
                row["image_urls"] = ""
                row["image_count"] = 0

            # Remove the complex images field
            del row["images"]
            flattened_data.append(row)

        # Write CSV
        fieldnames = flattened_data[0].keys()
        with open(output_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(flattened_data)

        self.stdout.write(f"Data exported to: {output_file}")
