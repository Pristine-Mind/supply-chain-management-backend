import hashlib
import logging
import random
from datetime import date, datetime, timedelta

import requests
from django.conf import settings
from django.core.cache import cache
from django.db import transaction
from django.db.models import Count, ExpressionWrapper, F, FloatField, Q, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.db.models.query import QuerySet
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from market.models import MarketplaceProduct
from market.serializers import MarketplaceProductSerializer
from user.models import UserProfile

# Cache TTLs (seconds) - can be overridden in Django settings
SEARCH_CACHE_TTL = getattr(settings, "PRODUCER_SEARCH_CACHE_TTL", 30)
LIST_CACHE_TTL = getattr(settings, "PRODUCER_LIST_CACHE_TTL", 15)

from .filters import (
    BrandFilter,
    CustomerFilter,
    MarketplaceProductFilter,
    OrderFilter,
    ProducerFilter,
    ProductFilter,
    SaleFilter,
)
from .models import (
    AuditLog,
    Brand,
    Category,
    City,
    Customer,
    DirectSale,
    LedgerEntry,
    MarketplaceProduct,
    Order,
    Payment,
    Producer,
    Product,
    PurchaseOrder,
    Sale,
    StockHistory,
    StockList,
    Subcategory,
    SubSubcategory,
)
from .serializers import (
    AuditLogSerializer,
    BrandSerializer,
    CategoryHierarchySerializer,
    CategorySerializer,
    CitySerializer,
    CustomerOrdersSerializer,
    CustomerSalesSerializer,
    CustomerSerializer,
    DirectSaleSerializer,
    KhaltiInitSerializer,
    KhaltiVerifySerializer,
    LedgerEntrySerializer,
    MarketplaceProductSerializer,
    OrderSerializer,
    PaymentSerializer,
    ProcurementRequestSerializer,
    ProcurementResponseSerializer,
    ProducerSerializer,
    ProductSerializer,
    ProductStockUpdateSerializer,
    PurchaseOrderSerializer,
    ReconciliationResponseSerializer,
    SaleSerializer,
    SalesRequestSerializer,
    SalesResponseSerializer,
    ShopQRSerializer,
    StockHistorySerializer,
    StockListSerializer,
    SubcategoryLightSerializer,
    SubcategorySerializer,
    SubSubcategoryLightSerializer,
    SubSubcategorySerializer,
)
from .supply_chain import SupplyChainService
from .utils import export_queryset_to_excel

logger = logging.getLogger(__name__)

from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated


class StockHistoryViewSet(viewsets.ModelViewSet):
    serializer_class = StockHistorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Only return active entries
        return StockHistory.objects.filter(is_active=True).order_by("-date")

    def perform_create(self, serializer):
        # Always set user from request
        serializer.save(user=self.request.user)

    def update(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return Response({"detail": "Updates not allowed. StockHistory entries are immutable."}, status=403)
        instance = self.get_object()
        instance._request = request
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return Response({"detail": "Updates not allowed. StockHistory entries are immutable."}, status=403)
        instance = self.get_object()
        instance._request = request
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            return Response({"detail": "Delete not allowed. Only superusers can delete."}, status=403)
        instance = self.get_object()
        instance._request = request
        # Soft-delete
        instance.is_active = False
        instance.save(update_fields=["is_active"])
        return Response(status=204)


class DailyProductStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")
        product_id = request.query_params.get("product")
        today = datetime.today().date()
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            except ValueError:
                return Response({"error": "Invalid start_date format. Use YYYY-MM-DD."}, status=400)
        else:
            start_date = today
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                return Response({"error": "Invalid end_date format. Use YYYY-MM-DD."}, status=400)
        else:
            end_date = start_date
        if end_date < start_date:
            return Response({"error": "end_date cannot be before start_date."}, status=400)
        # Filter products
        products = Product.objects.select_related(
            "brand", "category", "subcategory", "sub_subcategory", "producer", "user", "location"
        ).filter(user=user)
        if product_id:
            products = products.filter(id=product_id)
        export_type = request.query_params.get("export")
        result = []
        product_list = list(products)
        if export_type == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "Product Stats"
            headers = [
                "Product ID",
                "Product Name",
                "Producer Name",
                "SKU",
                "Category",
                "Price",
                "Cost Price",
                "Stock",
                "Reorder Level",
                "Date",
                "Stock In",
                "Stock Out",
                "Total Sales",
                "Total Sales Price",
                "Opening Stock",
                "Closing Stock",
            ]
            ws.append(headers)
            for product in product_list:
                product_info = {
                    "product_id": product.id,
                    "product_name": product.name,
                    "sku": product.sku,
                    "category": (
                        product.get_category_display() if hasattr(product, "get_category_display") else product.category
                    ),
                    "price": product.price,
                    "cost_price": product.cost_price,
                    "stock": product.stock,
                    "reorder_level": product.reorder_level,
                }
                # Prepare daily stats
                stats_by_day = {}
                opening_stock = None
                prev_stock_hist = (
                    StockHistory.objects.filter(product=product, user=user, is_active=True, date__lt=start_date)
                    .order_by("-date", "-id")
                    .first()
                )
                if prev_stock_hist and prev_stock_hist.stock_after is not None:
                    opening_stock = prev_stock_hist.stock_after
                else:
                    opening_stock = product.stock
                    all_hist = StockHistory.objects.filter(product=product, user=user, is_active=True, date__gte=start_date)
                    for hist in all_hist:
                        opening_stock -= hist.quantity_in - hist.quantity_out
                stock_histories = (
                    StockHistory.objects.filter(
                        product=product, user=user, is_active=True, date__gte=start_date, date__lte=end_date
                    )
                    .annotate(day=TruncDate("date"))
                    .values("day")
                    .annotate(
                        stock_in=Sum("quantity_in"), stock_out=Sum("quantity_out"), last_stock_after=Sum("stock_after")
                    )
                )
                sales = (
                    Sale.objects.filter(
                        order__product=product, user=user, sale_date__date__gte=start_date, sale_date__date__lte=end_date
                    )
                    .annotate(day=TruncDate("sale_date"))
                    .values("day")
                    .annotate(total_sales=Sum("quantity"), total_sales_price=Sum(F("quantity") * F("sale_price")))
                )
                for sh in stock_histories:
                    stats_by_day[sh["day"]] = {
                        "date": sh["day"],
                        "stock_in": sh["stock_in"] or 0,
                        "stock_out": sh["stock_out"] or 0,
                        "total_sales": 0,
                        "total_sales_price": 0,
                    }
                for sale in sales:
                    day = sale["day"]
                    if day not in stats_by_day:
                        stats_by_day[day] = {
                            "date": day,
                            "stock_in": 0,
                            "stock_out": 0,
                            "total_sales": sale["total_sales"] or 0,
                            "total_sales_price": sale["total_sales_price"] or 0,
                        }
                    else:
                        stats_by_day[day]["total_sales"] = sale["total_sales"] or 0
                        stats_by_day[day]["total_sales_price"] = sale["total_sales_price"] or 0
                num_days = (end_date - start_date).days + 1
                for i in range(num_days):
                    day = start_date + timedelta(days=i)
                    if day not in stats_by_day:
                        stats_by_day[day] = {
                            "date": day,
                            "stock_in": 0,
                            "stock_out": 0,
                            "total_sales": 0,
                            "total_sales_price": 0,
                        }
                sorted_days = sorted(stats_by_day.keys())
                prev_closing = opening_stock
                for day in sorted_days:
                    stats = stats_by_day[day]
                    stats["opening_stock"] = prev_closing
                    stats["closing_stock"] = prev_closing + stats["stock_in"] - stats["stock_out"]
                    prev_closing = stats["closing_stock"]
                stats_list = [stats_by_day[day] for day in sorted_days]
                for stats in stats_list:
                    ws.append(
                        [
                            product.id,
                            product.name,
                            product.user.get_full_name() or product.user.username if product.user else "N/A",
                            product.sku,
                            product.get_category_display() if hasattr(product, "get_category_display") else product.category,
                            product.price,
                            product.cost_price,
                            product.stock,
                            product.reorder_level,
                            stats["date"],
                            stats["stock_in"],
                            stats["stock_out"],
                            stats["total_sales"],
                            stats["total_sales_price"],
                            abs(stats["opening_stock"]),
                            abs(stats["closing_stock"]),
                        ]
                    )
            for i, col in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(i)].width = 15
            from io import BytesIO

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="product_stats.xlsx"'
            return response
        paginator = PageNumberPagination()
        paginator.page_size_query_param = "page_size"
        paginated_products = paginator.paginate_queryset(products, request)
        result = []
        for product in paginated_products:
            product_info = {
                "product_id": product.id,
                "product_name": product.name,
                "producer_name": product.user.get_full_name() or product.user.username if product.user else "N/A",
                "producer_id": product.user.id if product.user else None,
                "sku": product.sku,
                "category": product.get_category_display() if hasattr(product, "get_category_display") else product.category,
                "price": product.price,
                "cost_price": product.cost_price,
                "stock": product.stock,
                "reorder_level": product.reorder_level,
            }
            stats_by_day = {}
            prev_stock_hist = (
                StockHistory.objects.filter(product=product, user=user, is_active=True, date__lt=start_date)
                .order_by("-date", "-id")
                .first()
            )
            if prev_stock_hist and prev_stock_hist.stock_after is not None:
                opening_stock = prev_stock_hist.stock_after
                print(
                    f"[DEBUG] Product: {product.name}, Opening stock from last history before {start_date}: {opening_stock}"
                )
            else:
                opening_stock = 0
                print(
                    f"[WARNING] Product: {product.name}, No stock history before {start_date}. Defaulting opening stock to 0."
                )
            stock_histories = (
                StockHistory.objects.filter(
                    product=product, user=user, is_active=True, date__gte=start_date, date__lte=end_date
                )
                .annotate(day=TruncDate("date"))
                .values("day")
                .annotate(
                    stock_in=Sum("quantity_in"),
                    stock_out=Sum("quantity_out"),
                    last_stock_after=Sum("stock_after"),  # Not used for closing, just for completeness
                )
            )
            sales = (
                Sale.objects.filter(
                    order__product=product, user=user, sale_date__date__gte=start_date, sale_date__date__lte=end_date
                )
                .annotate(day=TruncDate("sale_date"))
                .values("day")
                .annotate(total_sales=Sum("quantity"), total_sales_price=Sum(F("quantity") * F("sale_price")))
            )
            for sh in stock_histories:
                stats_by_day[sh["day"]] = {
                    "date": sh["day"],
                    "stock_in": sh["stock_in"] or 0,
                    "stock_out": sh["stock_out"] or 0,
                    "total_sales": 0,
                    "total_sales_price": 0,
                }
            for sale in sales:
                day = sale["day"]
                if day not in stats_by_day:
                    stats_by_day[day] = {
                        "date": day,
                        "stock_in": 0,
                        "stock_out": 0,
                        "total_sales": sale["total_sales"] or 0,
                        "total_sales_price": sale["total_sales_price"] or 0,
                    }
                else:
                    stats_by_day[day]["total_sales"] = sale["total_sales"] or 0
                    stats_by_day[day]["total_sales_price"] = sale["total_sales_price"] or 0
            num_days = (end_date - start_date).days + 1
            for i in range(num_days):
                day = start_date + timedelta(days=i)
                if day not in stats_by_day:
                    stats_by_day[day] = {
                        "date": day,
                        "stock_in": 0,
                        "stock_out": 0,
                        "total_sales": 0,
                        "total_sales_price": 0,
                    }
            sorted_days = sorted(stats_by_day.keys())
            prev_closing = opening_stock
            for day in sorted_days:
                stats = stats_by_day[day]
                stats["opening_stock"] = prev_closing
                stats["closing_stock"] = prev_closing + stats["stock_in"] - stats["stock_out"]
                prev_closing = stats["closing_stock"]
            stats_list = [stats_by_day[day] for day in sorted_days]
            totals = {
                "stock_in": sum(s["stock_in"] for s in stats_list),
                "stock_out": sum(s["stock_out"] for s in stats_list),
                "total_sales": sum(s["total_sales"] for s in stats_list),
                "total_sales_price": sum(s["total_sales_price"] for s in stats_list),
                "opening_stock": abs(stats_list[0]["opening_stock"]) if stats_list else opening_stock,
                "closing_stock": abs(stats_list[-1]["closing_stock"]) if stats_list else opening_stock,
            }
            product_info.update(
                {
                    "stats": stats_list,
                    "totals": totals,
                }
            )
            result.append(product_info)
        return paginator.get_paginated_response(result)


class ProducerViewSet(viewsets.ModelViewSet):
    """
    A viewset for viewing and editing producer instances.
    """

    queryset = Producer.objects.select_related("user").all().order_by("-created_at")
    serializer_class = ProducerSerializer
    filterset_class = ProducerFilter
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Producer.objects.none()

        user_profile = getattr(user, "user_profile", None)
        if user_profile:
            return Producer.objects.select_related("user").filter(user__user_profile__shop_id=user_profile.shop_id)
        else:
            return Producer.objects.none()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        producer = serializer.save()
        return Response(self.get_serializer(producer).data, status=status.HTTP_201_CREATED)


class CustomerViewSet(viewsets.ModelViewSet):
    """
    A viewset for viewing and editing customer instances.
    """

    queryset = Customer.objects.select_related("user").all().order_by("-created_at")
    serializer_class = CustomerSerializer
    filterset_class = CustomerFilter
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Customer.objects.none()

        user_profile = getattr(user, "user_profile", None)
        if user_profile:
            return Customer.objects.select_related("user").filter(user__user_profile__shop_id=user_profile.shop_id)
        else:
            return Customer.objects.none()


class ProductViewSet(viewsets.ModelViewSet):
    """
    A viewset for viewing and editing product instances.
    """

    queryset = (
        Product.objects.select_related("brand", "category", "subcategory", "sub_subcategory", "producer", "user", "location")
        .prefetch_related("images")
        .all()
        .order_by("-created_at")
    )
    serializer_class = ProductSerializer
    filterset_class = ProductFilter
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Product.objects.none()

        user_profile = getattr(user, "user_profile", None)
        if user_profile:
            return (
                Product.objects.select_related(
                    "brand", "category", "subcategory", "sub_subcategory", "producer", "user", "location"
                )
                .prefetch_related("images")
                .filter(user__user_profile__shop_id=user_profile.shop_id)
            )
        else:
            return Product.objects.none()

    @action(detail=False, url_path="catgeories", methods=("get",), permission_classes=[AllowAny])
    def get_category(self, request, pk=None):
        return Response(
            [
                {
                    "key": key,
                    "value": value,
                }
                for key, value in Product.ProductCategory.choices
            ]
        )

    @action(detail=False, url_path="size-choices", methods=("get",), permission_classes=[AllowAny])
    def get_size_choices(self, request, pk=None):
        """Get available size choices"""
        return Response(
            [
                {
                    "key": key,
                    "value": value,
                }
                for key, value in Product.SizeChoices.choices
            ]
        )

    @action(detail=False, url_path="color-choices", methods=("get",), permission_classes=[AllowAny])
    def get_color_choices(self, request, pk=None):
        """Get available color choices"""
        return Response(
            [
                {
                    "key": key,
                    "value": value,
                }
                for key, value in Product.ColorChoices.choices
            ]
        )

    @action(
        detail=True,
        methods=["post"],
        url_path="update-stock",
        permission_classes=[IsAuthenticated],
        serializer_class=ProductStockUpdateSerializer,
    )
    def update_stock(self, request, pk=None):
        try:
            product = Product.objects.get(pk=pk)
        except Product.DoesNotExist:
            return Response({"detail": "Product not found."}, status=404)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        stock = serializer.validated_data["stock"]
        product.stock = stock
        product.updated_at = timezone.now()
        product.save(update_fields=["stock", "updated_at"])
        return Response(ProductSerializer(product).data)

    @action(detail=True, methods=["post"], url_path="push-to-marketplace", permission_classes=[IsAuthenticated])
    def push_to_marketplace(self, request, pk=None):
        """
        Push a product to the marketplace with optional custom settings

        Request body can contain:
        {
            "listed_price": 99.99,  // optional, defaults to product price
            "size": "M",            // optional, inherits from product if not provided
            "color": "RED",          // optional, inherits from product if not provided
            "additional_information": "Special marketplace info",  // optional
            // ... other marketplace-specific fields
        }
        """
        try:
            product = self.get_object()
        except Product.DoesNotExist:
            return Response({"error": "Product not found"}, status=status.HTTP_404_NOT_FOUND)

        # Check if marketplace product already exists
        if MarketplaceProduct.objects.filter(product=product).exists():
            return Response(
                {"error": f"Product '{product.name}' is already listed in the marketplace"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check if product is active
        if not product.is_active:
            return Response({"error": "Cannot push inactive product to marketplace"}, status=status.HTTP_400_BAD_REQUEST)

        # Prepare data for marketplace product creation
        marketplace_data = request.data.copy()
        marketplace_data["product_id"] = product.id

        # Use the same serializer as the create_from_product endpoint
        from .serializers import CreateMarketplaceProductFromProductSerializer

        serializer = CreateMarketplaceProductFromProductSerializer(data=marketplace_data)
        serializer.is_valid(raise_exception=True)

        try:
            marketplace_product = serializer.save()

            # Serialize the created marketplace product for response
            response_serializer = MarketplaceProductSerializer(marketplace_product, context={"request": request})

            return Response(
                {
                    "message": f"Product '{product.name}' has been successfully pushed to the marketplace",
                    "data": response_serializer.data,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response(
                {"error": f"Failed to push product to marketplace: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST
            )


class OrderViewSet(viewsets.ModelViewSet):
    """
    A viewset for viewing and editing order instances.
    """

    queryset = (
        Order.objects.select_related(
            "customer", "product", "product__brand", "product__category", "product__producer", "user"
        )
        .all()
        .order_by("-created_at")
    )
    serializer_class = OrderSerializer
    filterset_class = OrderFilter
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Order.objects.none()

        user_profile = getattr(user, "user_profile", None)
        if user_profile:
            return Order.objects.select_related(
                "customer", "product", "product__brand", "product__category", "product__producer", "user"
            ).filter(user__user_profile__shop_id=user_profile.shop_id)
        else:
            return Order.objects.none()

    @action(detail=True, methods=["post"])
    def update_status(self, request, pk=None):
        """
        Update the status of an order.
        Expected payload: {"status": "new_status"}
        Valid statuses: pending, approved, shipped, delivered, cancelled
        """
        order = self.get_object()
        new_status = request.data.get("status")

        if not new_status:
            return Response({"detail": "Status is required."}, status=status.HTTP_400_BAD_REQUEST)

        valid_statuses = [choice[0] for choice in Order.Status.choices]

        if new_status not in valid_statuses:
            return Response(
                {"detail": f"Invalid status. Must be one of: {', '.join(valid_statuses)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        order.status = new_status

        if new_status == Order.Status.DELIVERED and not order.delivery_date:
            order.delivery_date = timezone.now()

        order.save(update_fields=["status", "delivery_date", "updated_at"])

        if new_status == Order.Status.DELIVERED:
            product = order.product
            with transaction.atomic():
                product.refresh_from_db()
                product.stock = F("stock") - order.quantity
                product.save(update_fields=["stock"])

                StockHistory.objects.create(
                    product=product,
                    quantity_out=order.quantity,
                    notes=f"Stock out for order #{order.order_number}",
                    user=request.user,
                    stock_after=product.stock - order.quantity,
                )

        return Response(OrderSerializer(order).data, status=status.HTTP_200_OK)


class SaleViewSet(viewsets.ModelViewSet):
    """
    A viewset for viewing and editing sale instances.
    """

    queryset = (
        Sale.objects.select_related(
            "order",
            "order__customer",
            "order__product",
            "order__product__brand",
            "order__product__category",
            "order__product__producer",
            "user",
            "payment",
        )
        .all()
        .order_by("-created_at")
    )
    serializer_class = SaleSerializer
    filterset_class = SaleFilter
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return Sale.objects.none()

        user_profile = getattr(user, "user_profile", None)
        if user_profile:
            return Sale.objects.select_related(
                "order",
                "order__customer",
                "order__product",
                "order__product__brand",
                "order__product__category",
                "order__product__producer",
                "user",
                "payment",
            ).filter(user__user_profile__shop_id=user_profile.shop_id)
        else:
            return Sale.objects.none()


class DashboardAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if not user.is_authenticated:
            return Response({"detail": "Authentication required"}, status=401)

        user_profile = getattr(user, "user_profile", None)
        if not user_profile:
            return Response({"detail": "User profile not found"}, status=404)

        shop_id = user_profile.shop_id
        current_year = timezone.now().year

        total_products = Product.objects.filter(user__user_profile__shop_id=shop_id).distinct().count() or 0
        total_orders = Order.objects.filter(user__user_profile__shop_id=shop_id).count()
        total_sales = Sale.objects.filter(user__user_profile__shop_id=shop_id).count()
        total_customers = Customer.objects.filter(user__user_profile__shop_id=shop_id).count()
        pending_orders = Order.objects.filter(
            status__in=[Order.Status.PENDING, Order.Status.APPROVED], user__user_profile__shop_id=shop_id
        ).count()
        total_revenue = (
            Sale.objects.filter(user__user_profile__shop_id=shop_id).aggregate(total_revenue=Sum("sale_price"))[
                "total_revenue"
            ]
            or 0
        )

        sales_trends = (
            Sale.objects.filter(user__user_profile__shop_id=shop_id, sale_date__year=current_year)
            .annotate(month=TruncMonth("sale_date"))
            .values("month")
            .annotate(total_sales=Sum("sale_price"))
            .order_by("month")
        )

        data = {
            "totalProducts": total_products,
            "totalOrders": total_orders,
            "totalSales": total_sales,
            "totalCustomers": total_customers,
            "pendingOrders": pending_orders,
            "totalRevenue": total_revenue,
            "salesTrends": [{"month": sale["month"].strftime("%B"), "value": sale["total_sales"]} for sale in sales_trends],
        }

        return Response(data)


class UserInfoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        user_profile = None
        try:
            user_profile = UserProfile.objects.get(user=user)
            has_access_to_marketplace = user_profile.has_access_to_marketplace
            b2b_verified = getattr(user_profile, "b2b_verified", False)
        except UserProfile.DoesNotExist:
            has_access_to_marketplace = False
            b2b_verified = False

        return Response(
            {
                "username": user.username,
                "id": user.id,
                "has_access_to_marketplace": has_access_to_marketplace,
                "business_type": user_profile.business_type if user_profile else None,
                "role": user_profile.role.code if user_profile and user_profile.role else None,
                "email": user.email,
                "b2b_verified": b2b_verified,
            }
        )


class TopSalesCustomersView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        current_year = timezone.now().year
        user = request.user
        if not user.is_authenticated:
            return Response({"detail": "Authentication required"}, status=401)

        user_profile = getattr(user, "user_profile", None)
        if not user_profile:
            return Response({"detail": "User profile not found"}, status=404)

        shop_id = user_profile.shop_id

        top_sales_customers = (
            Sale.objects.filter(user__user_profile__shop_id=shop_id, sale_date__year=current_year)
            .annotate(total_sales_amount=ExpressionWrapper(F("sale_price") * F("quantity"), output_field=FloatField()))
            .values("order__customer__id", "order__customer__name")
            .annotate(total_sales=Sum("total_sales_amount"), name=F("order__customer__name"), id=F("order__customer__id"))
            .order_by("-total_sales")
        )
        sales_serializer = CustomerSalesSerializer(top_sales_customers, many=True)
        return Response(data=sales_serializer.data, status=status.HTTP_200_OK)


class TopOrdersCustomersView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        current_year = timezone.now().year
        user = request.user
        if not user.is_authenticated:
            return Response({"detail": "Authentication required"}, status=401)

        user_profile = getattr(user, "user_profile", None)
        if not user_profile:
            return Response({"detail": "User profile not found"}, status=404)

        shop_id = user_profile.shop_id

        top_orders_customers = (
            Customer.objects.filter(user__user_profile__shop_id=shop_id, order__order_date__year=current_year)
            .annotate(total_orders=Count("order"))
            .order_by("-total_orders")[:10]
        )

        orders_serializer = CustomerOrdersSerializer(top_orders_customers, many=True)
        return Response(orders_serializer.data, status=status.HTTP_200_OK)


class StockListView(viewsets.ModelViewSet):
    queryset = StockList.objects.filter(is_pushed_to_marketplace=False).distinct()
    serializer_class = StockListSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return StockList.objects.none()

        user_profile = getattr(user, "user_profile", None)
        if user_profile:
            return self.queryset.filter(user__user_profile__shop_id=user_profile.shop_id)
        else:
            return StockList.objects.none()

    @action(detail=True, methods=["post"], url_path="push-to-marketplace")
    def push_to_marketplace(self, request, pk=None):
        try:
            stock_item = self.get_object()
        except StockList.DoesNotExist:
            return Response({"error": "StockList item not found."}, status=status.HTTP_404_NOT_FOUND)

        if MarketplaceProduct.objects.filter(product=stock_item.product).exists():
            return Response(
                {"error": f"Product '{stock_item.product.name}' is already listed in the marketplace."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        MarketplaceProduct.objects.create(
            product=stock_item.product,
            listed_price=stock_item.product.price,
            is_available=True,
            bid_end_date=timezone.now() + timedelta(days=7),
        )

        stock_item.is_pushed_to_marketplace = True
        stock_item.moved_date = timezone.now()
        stock_item.save(update_fields=["is_pushed_to_marketplace"])
        return Response(
            {"message": f"Product '{stock_item.product.name}' has been successfully pushed to the marketplace."},
            status=status.HTTP_200_OK,
        )


class MarketplaceProductViewSet(viewsets.ModelViewSet):
    serializer_class = MarketplaceProductSerializer
    filterset_class = MarketplaceProductFilter

    def get_queryset(self):
        return (
            MarketplaceProduct.objects.filter(is_available=True)
            .select_related(
                "product",
                "product__user",
                "product__user__user_profile",
                "product__brand",
                "product__category",
                "product__subcategory",
                "product__sub_subcategory",
                "product__producer",
                "product__location",
            )
            .prefetch_related("bulk_price_tiers", "variants", "reviews", "product__images")
            .order_by("-listed_date")
            .distinct()
        )

    def list(self, request, *args, **kwargs):
        """List marketplace products but randomize order within the first N results.

        This takes the filtered queryset (so filtering backends still apply),
        limits to `first_n` products (default 50) ordered by listed_date,
        shuffles them in-memory, then paginates the shuffled list.
        """
        # Apply filtering backends (search, filterset, etc.)
        qs = self.filter_queryset(self.get_queryset())

        # Determine how many top products to randomize from
        try:
            first_n = int(request.query_params.get("first_n", 50))
        except (ValueError, TypeError):
            first_n = 50

        # Take a candidate pool (larger than first_n) to allow picking from
        # multiple categories, then build a category-balanced list by
        # round-robin selecting from each category in random order.
        candidate_factor = 3
        max_candidate_cap = 500
        candidate_size = min(max_candidate_cap, max(first_n * candidate_factor, first_n))
        candidates = list(qs[:candidate_size])

        # Group candidates by category (use category id or name where available)
        from collections import defaultdict

        buckets = defaultdict(list)
        for mp in candidates:
            try:
                cat = mp.product.category
            except Exception:
                cat = None
            # Use id if it's a model, else use the value directly
            cat_key = getattr(cat, "id", None) if getattr(cat, "id", None) is not None else cat
            buckets[cat_key].append(mp)

        # Randomize bucket order so categories are interleaved unpredictably
        category_keys = list(buckets.keys())
        random.shuffle(category_keys)

        limited_qs = []
        # Round-robin pick one from each bucket until we reach first_n
        while len(limited_qs) < first_n and any(buckets[k] for k in category_keys):
            for k in list(category_keys):
                if not buckets[k]:
                    continue
                # Pop from front to keep higher-ranked items earlier for each category
                limited_qs.append(buckets[k].pop(0))
                if len(limited_qs) >= first_n:
                    break

        # If still short, append remaining candidates in order
        if len(limited_qs) < first_n:
            remaining = [mp for k in category_keys for mp in buckets[k]]
            for mp in remaining:
                if len(limited_qs) >= first_n:
                    break
                limited_qs.append(mp)

        # Build the final ordered list: randomized first `first_n`, then the rest
        try:
            selected_ids = [int(mp.id) for mp in limited_qs]
        except Exception:
            selected_ids = []

        # Get remaining products (preserve original queryset ordering)
        try:
            remaining_qs = list(qs.exclude(id__in=selected_ids)) if selected_ids else list(qs)
        except Exception:
            # Fallback: use candidates not already selected
            remaining_qs = [mp for mp in list(qs) if getattr(mp, "id", None) not in selected_ids]

        final_list = limited_qs + remaining_qs

        # Use DRF paginator on the combined list
        try:
            user_part = str(request.user.id) if getattr(request, "user", None) and request.user.is_authenticated else "anon"
        except Exception:
            user_part = "anon"
        cache_key = (
            "producer:list:" + hashlib.sha256((request.get_full_path() + ":" + user_part).encode("utf-8")).hexdigest()
        )
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached)

        paginator = self.paginator or PageNumberPagination()
        # Allow page_size via query param if provided
        page_size = request.query_params.get("page_size")
        if page_size:
            try:
                paginator.page_size = int(page_size)
            except (ValueError, TypeError):
                pass

        page = paginator.paginate_queryset(final_list, request, view=self)
        serializer = self.get_serializer(page, many=True, context={"request": request})
        response = paginator.get_paginated_response(serializer.data)

        # Cache the response data (dict) for a short TTL
        try:
            cache.set(cache_key, response.data, LIST_CACHE_TTL)
        except Exception:
            pass

        return response

    @action(detail=False, methods=["get"], url_path="search", permission_classes=[AllowAny])
    def search(self, request):
        from haystack.query import SQ, SearchQuerySet
        from rest_framework.pagination import PageNumberPagination

        phrase = request.query_params.get("keyword") or request.query_params.get("q")
        if phrase is None:
            return Response({"detail": "Must include a `keyword` or `q` query parameter."}, status=400)

        phrase = phrase.strip()
        if not phrase:
            return Response({"detail": "Empty `keyword` provided."}, status=400)

        category = request.query_params.get("category")
        sub_category = request.query_params.get("sub_category") or request.query_params.get("subcategory")
        sub_subcategory = request.query_params.get("sub_subcategory")

        # New filters
        brand = request.query_params.get("brand")
        sku = request.query_params.get("sku")
        size = request.query_params.get("size")
        color = request.query_params.get("color")
        min_price = request.query_params.get("min_price")
        max_price = request.query_params.get("max_price")

        # Cache search results for identical queries for a short TTL
        try:
            user_part = str(request.user.id) if getattr(request, "user", None) and request.user.is_authenticated else "anon"
        except Exception:
            user_part = "anon"
        cache_key = (
            "producer:search:" + hashlib.sha256((request.get_full_path() + ":" + user_part).encode("utf-8")).hexdigest()
        )
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached)

        sqs = SearchQuerySet().models(MarketplaceProduct)

        try:
            combined = (
                SQ(name__startswith=phrase)
                | SQ(description__contains=phrase)
                | SQ(category__contains=phrase)
                | SQ(subcategory__contains=phrase)
                | SQ(sub_subcategory__contains=phrase)
                | SQ(text__contains=phrase)
                | SQ(search_tags__contains=phrase)
                | SQ(brand__contains=phrase)
                | SQ(sku__contains=phrase)
            )
            sqs = sqs.filter(combined).order_by("-_score")
        except Exception:
            try:
                sqs = sqs.auto_query(phrase).order_by("-_score")
            except Exception:
                sqs = sqs.filter(content=phrase)

        def _id_to_name(val):
            try:
                int_val = int(val)
            except Exception:
                return val
            from producer.models import Category, Subcategory, SubSubcategory

            # Try category, then subcategory, then sub_subcategory
            try:
                c = Category.objects.filter(id=int_val).first()
                if c:
                    return c.name
            except Exception:
                pass
            try:
                s = Subcategory.objects.filter(id=int_val).first()
                if s:
                    return s.name
            except Exception:
                pass
            try:
                ss = SubSubcategory.objects.filter(id=int_val).first()
                if ss:
                    return ss.name
            except Exception:
                pass
            return val

        if category:
            name_val = _id_to_name(category)
            try:
                sqs = sqs.filter(SQ(category=name_val) | SQ(subcategory=name_val) | SQ(sub_subcategory=name_val))
            except Exception:
                sqs = sqs.filter(
                    SQ(category__contains=name_val)
                    | SQ(subcategory__contains=name_val)
                    | SQ(sub_subcategory__contains=name_val)
                )
        else:

            def _apply_field_filter(sqs_obj, field_name, value):
                if not value:
                    return sqs_obj
                try:
                    int_val = int(value)
                except Exception:
                    int_val = None

                if int_val:
                    from producer.models import Category, Subcategory, SubSubcategory

                    try:
                        if field_name == "subcategory":
                            sub = Subcategory.objects.get(id=int_val)
                            return sqs_obj.filter(subcategory=sub.name)
                        if field_name == "sub_subcategory":
                            ssub = SubSubcategory.objects.get(id=int_val)
                            return sqs_obj.filter(sub_subcategory=ssub.name)
                    except Exception:
                        return sqs_obj

                return sqs_obj.filter(**{field_name: value})

            sqs = _apply_field_filter(sqs, "subcategory", sub_category)
            sqs = _apply_field_filter(sqs, "sub_subcategory", sub_subcategory)

        # Apply additional filters
        if brand:
            sqs = sqs.filter(brand__icontains=brand)
        if sku:
            sqs = sqs.filter(sku__icontains=sku)
        if size:
            sqs = sqs.filter(size__iexact=size)
        if color:
            sqs = sqs.filter(color__iexact=color)

        if min_price:
            try:
                sqs = sqs.filter(listed_price__gte=float(min_price))
            except ValueError:
                pass

        if max_price:
            try:
                sqs = sqs.filter(listed_price__lte=float(max_price))
            except ValueError:
                pass

        # Optimize: Paginate the SearchQuerySet directly to avoid fetching all objects from DB
        paginator = PageNumberPagination()
        paginator.page_size = int(request.query_params.get("page_size", 20))

        # This returns a list of SearchResult objects for the current page
        page_results = paginator.paginate_queryset(sqs, request, view=self)

        # Resolve objects only for the current page
        results = [r.object for r in page_results if getattr(r, "object", None) is not None]

        serializer = self.get_serializer(results, many=True, context={"request": request})
        response = paginator.get_paginated_response(serializer.data)

        # Cache the response payload for a short TTL
        try:
            cache.set(cache_key, response.data, SEARCH_CACHE_TTL)
        except Exception:
            pass

        return response

    @action(detail=False, url_path="size-choices", methods=("get",), permission_classes=[AllowAny])
    def get_size_choices(self, request, pk=None):
        """Get available size choices for marketplace products"""
        return Response(
            [
                {
                    "key": key,
                    "value": value,
                }
                for key, value in MarketplaceProduct.SizeChoices.choices
            ]
        )

    @action(detail=False, url_path="color-choices", methods=("get",), permission_classes=[AllowAny])
    def get_color_choices(self, request, pk=None):
        """Get available color choices for marketplace products"""
        return Response(
            [
                {
                    "key": key,
                    "value": value,
                }
                for key, value in MarketplaceProduct.ColorChoices.choices
            ]
        )

    @action(detail=False, url_path="filter-options", methods=("get",), permission_classes=[AllowAny])
    def get_filter_options(self, request, pk=None):
        """Get all available filter options including sizes and colors"""
        return Response(
            {
                "sizes": [{"key": key, "value": value} for key, value in MarketplaceProduct.SizeChoices.choices],
                "colors": [{"key": key, "value": value} for key, value in MarketplaceProduct.ColorChoices.choices],
            }
        )

    @action(detail=False, url_path="create-from-product", methods=["post"], permission_classes=[IsAuthenticated])
    def create_from_product(self, request):
        """
        Create a marketplace product from an existing product

        Request body should contain:
        {
            "product_id": 123,
            "listed_price": 99.99,  // optional, defaults to product price
            "size": "M",            // optional, inherits from product if not provided
            "color": "RED",          // optional, inherits from product if not provided
            "additional_information": "Special marketplace info",  // optional
            // ... other marketplace-specific fields
        }
        """
        from .serializers import CreateMarketplaceProductFromProductSerializer

        serializer = CreateMarketplaceProductFromProductSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Get the product to verify user permissions
        product_id = serializer.validated_data["product_id"]
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response({"error": "Product not found"}, status=status.HTTP_404_NOT_FOUND)

        # Check if user has permission to create marketplace product for this product
        user = request.user
        user_profile = getattr(user, "user_profile", None)

        if not user.is_staff and not user.is_superuser:
            # Regular users can only create marketplace products for their own products
            if user_profile:
                if product.user.user_profile.shop_id != user_profile.shop_id:
                    return Response(
                        {"error": "You don't have permission to create marketplace product for this product"},
                        status=status.HTTP_403_FORBIDDEN,
                    )
            else:
                if product.user != user:
                    return Response(
                        {"error": "You can only create marketplace products for your own products"},
                        status=status.HTTP_403_FORBIDDEN,
                    )

        # Create the marketplace product
        try:
            marketplace_product = serializer.save()

            # Serialize the created marketplace product for response
            response_serializer = MarketplaceProductSerializer(marketplace_product, context={"request": request})

            return Response(
                {
                    "message": f"Marketplace product created successfully for '{product.name}'",
                    "data": response_serializer.data,
                },
                status=status.HTTP_201_CREATED,
            )

        except Exception as e:
            return Response({"error": f"Failed to create marketplace product: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["get"], permission_classes=[IsAuthenticated])
    def b2b_pricing(self, request, pk=None):
        """Get B2B pricing information for authenticated business users"""
        product = self.get_object()
        user = request.user
        quantity = int(request.query_params.get("quantity", 1))

        if not user.is_authenticated:
            return Response({"detail": "Authentication required"}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            profile = getattr(user, "user_profile", None)
            if not profile or not getattr(profile, "is_b2b_eligible", False):
                return Response({"detail": "B2B verification required"}, status=status.HTTP_403_FORBIDDEN)
        except AttributeError:
            return Response({"detail": "B2B verification required"}, status=status.HTTP_403_FORBIDDEN)

        from .services import B2BPricingService

        pricing_info = B2BPricingService.get_b2b_pricing_for_product(product, user, quantity)

        return Response(pricing_info)

    @action(detail=True, methods=["get"], permission_classes=[IsAuthenticated])
    def bulk_pricing_breakdown(self, request, pk=None):
        """Get pricing breakdown for different quantities to show bulk discounts"""
        product = self.get_object()
        user = request.user
        max_quantity = int(request.query_params.get("max_quantity", 100))

        from .services import B2BPricingService

        breakdown = B2BPricingService.calculate_bulk_discount_breakdown(product, user, max_quantity)

        return Response({"product_id": product.id, "max_quantity": max_quantity, "pricing_tiers": breakdown})

    @action(detail=False, methods=["post"], permission_classes=[IsAuthenticated])
    def calculate_order_pricing(self, request):
        """Calculate pricing for multiple items in an order"""
        user = request.user
        items_data = request.data.get("items", [])

        if not items_data:
            return Response({"error": "No items provided"}, status=status.HTTP_400_BAD_REQUEST)

        items = []
        for item_data in items_data:
            try:
                product = MarketplaceProduct.objects.get(id=item_data["product_id"])
                items.append({"product": product, "quantity": item_data["quantity"]})
            except (MarketplaceProduct.DoesNotExist, KeyError):
                return Response({"error": f"Invalid product or quantity data"}, status=status.HTTP_400_BAD_REQUEST)

        from .services import B2BPricingService

        order_pricing = B2BPricingService.calculate_order_pricing(user, items)

        # Convert line items for serialization
        order_pricing["line_items"] = [
            {
                "product_id": item["product"].id,
                "product_name": item["product"].product.name,
                "quantity": item["quantity"],
                "unit_price": float(item["unit_price"]),
                "regular_price": float(item["regular_price"]),
                "line_total": float(item["line_total"]),
                "additional_discount": float(item["additional_discount"]),
                "is_b2b_price": item["is_b2b_price"],
                "savings": float(item["savings"]),
            }
            for item in order_pricing["line_items"]
        ]

        # Convert decimal fields
        order_pricing["subtotal"] = float(order_pricing["subtotal"])
        order_pricing["total_savings"] = float(order_pricing["total_savings"])

        return Response(order_pricing)


class MarketplaceUserRecommendedProductViewSet(viewsets.ModelViewSet):
    serializer_class = MarketplaceProductSerializer
    filterset_class = MarketplaceProductFilter
    # permission_classes = [IsAuthenticated]

    def get_queryset(self) -> QuerySet:
        # user = self.request.user
        # try:
        #     UserProfile.objects.get(user=user)
        #     # location = user_profile.location
        # except UserProfile.DoesNotExist:
        #     return MarketplaceProduct.objects.none()

        # if not location:
        #     return MarketplaceProduct.objects.none()

        return MarketplaceProduct.objects.filter(
            is_available=True,
            # bid_end_date__gte=timezone.now(),
            # product__location=location,
        ).order_by("-listed_date")


class StatsAPIView(APIView):
    """
    API to provide sales statistics with optional filtering.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user
        filter_params = self.get_filter_params(request)
        sales_query = self.get_filtered_sales(filter_params, user)
        data = self.get_sales_statistics(sales_query)
        return Response(data, status=status.HTTP_200_OK)

    def get_filter_params(self, request):
        """Get and parse filter parameters from the request."""
        location = request.query_params.get("location")
        category = request.query_params.get("category")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        parsed_start_date, parsed_end_date = None, None
        if start_date and end_date:
            try:
                parsed_start_date = timezone.datetime.strptime(start_date, "%Y-%m-%d")
                parsed_end_date = timezone.datetime.strptime(end_date, "%Y-%m-%d")
            except ValueError:
                return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        return {"location": location, "category": category, "start_date": parsed_start_date, "end_date": parsed_end_date}

    def get_filtered_sales(self, filter_params, user):
        """Filter the sales query based on the provided parameters."""
        if not user.is_authenticated:
            return Response({"detail": "Authentication required"}, status=401)

        user_profile = getattr(user, "user_profile", None)
        if not user_profile:
            return Response({"detail": "User profile not found"}, status=404)

        # Filter data based on the shop_id of the user's profile
        shop_id = user_profile.shop_id
        sales_query = Sale.objects.filter(user__user_profile__shop_id=shop_id)

        if filter_params["location"]:
            sales_query = sales_query.filter(order__customer__city=filter_params["location"])
        if filter_params["category"]:
            sales_query = sales_query.filter(order__product__category=filter_params["category"])
        if filter_params["start_date"] and filter_params["end_date"]:
            sales_query = sales_query.filter(
                sale_date__gte=filter_params["start_date"], sale_date__lte=filter_params["end_date"]
            )

        return sales_query

    def get_sales_statistics(self, sales_query):
        """Calculate sales statistics."""
        total_products_sold = sales_query.aggregate(total=Sum("quantity"))["total"] or 0
        total_revenue = sales_query.aggregate(revenue=Sum("sale_price"))["revenue"] or 0
        top_customers = (
            sales_query.values("order__customer__name", "order__customer__billing_address")
            .annotate(total_spent=Sum("sale_price"))
            .order_by("-total_spent")[:5]
        )
        top_products = (
            sales_query.values("order__product__name").annotate(total_sold=Sum("quantity")).order_by("-total_sold")[:5]
        )
        top_categories = (
            sales_query.values("order__product__category").annotate(total_sold=Sum("quantity")).order_by("-total_sold")[:5]
        )
        monthly_sales = (
            sales_query.annotate(month=TruncMonth("sale_date"))
            .values("month")
            .annotate(total_sold=Sum("quantity"))
            .order_by("month")
        )

        return {
            "total_products_sold": total_products_sold,
            "total_revenue": total_revenue,
            "top_customers": list(top_customers),
            "top_products": list(top_products),
            "top_categories": list(top_categories),
            "monthly_sales": list(monthly_sales),
        }


class CityListView(APIView):
    def get(self, request):
        cities = City.objects.all()
        serializer = CitySerializer(cities, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def withdraw_product(request, product_id):
    try:
        product = MarketplaceProduct.objects.get(pk=product_id, product__user=request.user)

        if product.bid_end_date and product.bid_end_date < timezone.now():
            return Response(
                {"error": "Cannot withdraw product. The bidding period has ended."}, status=status.HTTP_400_BAD_REQUEST
            )

        product.is_available = False
        product.save()

        return Response({"message": "Product withdrawn successfully."}, status=status.HTTP_200_OK)

    except MarketplaceProduct.DoesNotExist:
        return Response(
            {"error": "Product not found or you do not have permission to withdraw this product."},
            status=status.HTTP_404_NOT_FOUND,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_producers_to_excel(request):
    field_names = [
        "name",
        "contact",
        "email",
        "address",
        "registration_number",
        "created_at",
        "updated_at",
    ]
    user = request.user
    if not user.is_authenticated:
        return

    user_profile = getattr(user, "user_profile", None)
    if user_profile:
        queryset = Producer.objects.filter(user__user_profile__shop_id=user_profile.shop_id)
    wb = export_queryset_to_excel(queryset, field_names)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=producers.xlsx"
    wb.save(response)
    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_customers_to_excel(request):
    field_names = [
        "name",
        "customer_type",
        "contact",
        "email",
        "billing_address",
        "shipping_address",
        "credit_limit",
        "current_balance",
        "created_at",
        "updated_at",
    ]
    user = request.user
    if not user.is_authenticated:
        return

    user_profile = getattr(user, "user_profile", None)
    if user_profile:
        queryset = Customer.objects.filter(user__user_profile__shop_id=user_profile.shop_id)
    wb = export_queryset_to_excel(queryset, field_names)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=customers.xlsx"
    wb.save(response)
    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_products_to_excel(request):
    field_names = [
        "name",
        "get_category_display",
        "description",
        "sku",
        "price",
        "cost_price",
        "stock",
        "reorder_level",
        "is_active",
        "created_at",
        "updated_at",
    ]
    headers = [
        "Name",
        "Category",
        "Description",
        "SKU",
        "Price",
        "Cost Price",
        "Stock",
        "Reorder Level",
        "Is Active",
        "Created At",
        "Updated At",
    ]

    user = request.user
    if not user.is_authenticated:
        return

    user_profile = getattr(user, "user_profile", None)
    if user_profile:
        queryset = Product.objects.filter(user__user_profile__shop_id=user_profile.shop_id)
    wb = export_queryset_to_excel(queryset, field_names, headers)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=products.xlsx"
    wb.save(response)
    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_orders_to_excel(request):
    field_names = [
        "order_number",
        "customer_name",
        "product_name",
        "quantity",
        "get_status_display",
        "order_date",
        "delivery_date",
        "total_price",
        "created_at",
        "updated_at",
    ]
    headers = [
        "Order Number",
        "Customer",
        "Product",
        "Quantity",
        "Status",
        "Order Date",
        "Delivery Date",
        "Total Price",
        "Created At",
        "Updated At",
    ]

    user = request.user
    if not user.is_authenticated:
        return Response({"error": "Authentication required"}, status=status.HTTP_401_UNAUTHORIZED)

    start_date = request.query_params.get("start_date")
    end_date = request.query_params.get("end_date")
    product_id = request.query_params.get("product_id")

    user_profile = getattr(user, "user_profile", None)
    if not user_profile:
        return Response({"error": "User profile not found"}, status=status.HTTP_400_BAD_REQUEST)

    queryset = Order.objects.filter(user__user_profile__shop_id=user_profile.shop_id).select_related("customer", "product")
    if start_date:
        try:
            start_date = timezone.datetime.strptime(start_date, "%Y-%m-%d").date()
            queryset = queryset.filter(order_date__gte=start_date)
        except ValueError:
            return Response({"error": "Invalid start_date format. Use YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)

    if end_date:
        try:
            end_date = timezone.datetime.strptime(end_date, "%Y-%m-%d").date()
            end_date = end_date + timezone.timedelta(days=1)
            queryset = queryset.filter(order_date__lt=end_date)
        except ValueError:
            return Response({"error": "Invalid end_date format. Use YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)

    if product_id:
        try:
            product_id = int(product_id)
            queryset = queryset.filter(product_id=product_id)
        except (ValueError, TypeError):
            return Response({"error": "Invalid product_id"}, status=status.HTTP_400_BAD_REQUEST)

    queryset = queryset.annotate(customer_name=F("customer__name"), product_name=F("product__name"))
    wb = export_queryset_to_excel(queryset, field_names, headers)

    filename = "orders"
    if start_date or end_date:
        start_str = start_date.strftime("%Y%m%d") if start_date else "start"
        end_str = (end_date - timezone.timedelta(days=1)).strftime("%Y%m%d") if end_date else "end"
        filename += f"_{start_str}_to_{end_str}"
    if product_id:
        filename += f"_product_{product_id}"
    filename += ".xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_sales_to_excel(request):
    field_names = [
        "order_number",
        "product_name",
        "quantity",
        "sale_price",
        "sale_date",
        "get_payment_status_display",
        "payment_due_date",
        "created_at",
        "updated_at",
    ]
    headers = [
        "Order Number",
        "Product",
        "Quantity",
        "Sale Price",
        "Sale Date",
        "Payment Status",
        "Payment Due Date",
        "Created At",
        "Updated At",
    ]

    user = request.user
    if not user.is_authenticated:
        return Response({"error": "Authentication required"}, status=status.HTTP_401_UNAUTHORIZED)

    # Get query parameters
    start_date = request.query_params.get("start_date")
    end_date = request.query_params.get("end_date")
    product_id = request.query_params.get("product_id")

    user_profile = getattr(user, "user_profile", None)
    if not user_profile:
        return Response({"error": "User profile not found"}, status=status.HTTP_400_BAD_REQUEST)

    # Base queryset with shop filtering
    queryset = Sale.objects.filter(user__user_profile__shop_id=user_profile.shop_id).select_related("order__product")

    # Apply date filters
    if start_date:
        try:
            start_date = timezone.datetime.strptime(start_date, "%Y-%m-%d").date()
            queryset = queryset.filter(sale_date__gte=start_date)
        except ValueError:
            return Response({"error": "Invalid start_date format. Use YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)

    if end_date:
        try:
            end_date = timezone.datetime.strptime(end_date, "%Y-%m-%d").date()
            # Include the entire end date by adding 1 day
            end_date = end_date + timezone.timedelta(days=1)
            queryset = queryset.filter(sale_date__lt=end_date)
        except ValueError:
            return Response({"error": "Invalid end_date format. Use YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)

    # Apply product filter
    if product_id:
        try:
            product_id = int(product_id)
            queryset = queryset.filter(order__product_id=product_id)
        except (ValueError, TypeError):
            return Response({"error": "Invalid product_id"}, status=status.HTTP_400_BAD_REQUEST)

    # Annotate with order number and product name
    queryset = queryset.annotate(order_number=F("order__order_number"), product_name=F("order__product__name"))

    # Export to Excel
    wb = export_queryset_to_excel(queryset, field_names, headers)

    # Create response with filename including date range if applicable
    filename = "sales"
    if start_date or end_date:
        start_str = start_date.strftime("%Y%m%d") if start_date else "start"
        end_str = (end_date - timezone.timedelta(days=1)).strftime("%Y%m%d") if end_date else "end"
        filename += f"_{start_str}_to_{end_str}"
    if product_id:
        filename += f"_product_{product_id}"
    filename += ".xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response


class LedgerEntryViewSet(viewsets.ModelViewSet):
    queryset = LedgerEntry.objects.all()
    serializer_class = LedgerEntrySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)


class AuditLogViewSet(viewsets.ModelViewSet):
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return self.queryset.filter(user=self.request.user)


@api_view(["POST"])
def procurement_view(request):
    serializer = ProcurementRequestSerializer(data=request.data)
    if serializer.is_valid():
        service = SupplyChainService(request.user)
        try:
            order = service.procurement_process(
                producer_id=serializer.validated_data["producer_id"],
                product_id=serializer.validated_data["product_id"],
                quantity=serializer.validated_data["quantity"],
                unit_cost=serializer.validated_data["unit_cost"],
            )
            response_serializer = ProcurementResponseSerializer(order)
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
def sales_view(request):
    serializer = SalesRequestSerializer(data=request.data)
    if serializer.is_valid():
        service = SupplyChainService(request.user)
        try:
            sale = service.sales_process(
                customer_id=serializer.validated_data["customer_id"],
                product_id=serializer.validated_data["product_id"],
                quantity=serializer.validated_data["quantity"],
                selling_price=serializer.validated_data["selling_price"],
            )
            response_serializer = SalesResponseSerializer(sale)
            return Response(response_serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
def reconciliation_view(request):
    service = SupplyChainService(request.user)
    try:
        result = service.reconciliation_process()
        serializer = ReconciliationResponseSerializer(result)
        return Response(serializer.data)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def stats_dashboard(request):
    user = request.user

    user_date = request.query_params.get("date", None)
    if user_date:
        print(user_date)
        try:
            date = datetime.strptime(user_date, "%Y-%m-%d")
            print(f"date:{date}")
        except ValueError:
            return Response({"error": "Invalid date format. Please use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
    else:
        date = datetime.today().date()
    start_of_week = date - timedelta(days=date.weekday())
    print(f"start_of_week: {start_of_week}")
    start_of_month = date.replace(day=1)

    producer_orders = (
        Order.objects.filter(user=user, order_date__gte=start_of_week)
        .select_related("product", "customer")
        .order_by("-created_at")[:10]
    )
    customer_sales = (
        Sale.objects.filter(user=user, sale_date__gte=start_of_week)
        .select_related("order__product")
        .order_by("-created_at")[:10]
    )

    producer_transactions = [
        {
            "name": order.product.producer.name if order.product and order.product.producer else "Unknown",
            "detail": f"{order.quantity} units of {order.product.name}",
            "date": order.order_date.date(),
        }
        for order in producer_orders
    ]

    customer_transactions = [
        {
            "name": sale.order.customer.name if sale.order.customer else "Unknown",
            "detail": f"{sale.quantity} units of {sale.order.product.name}",
            "date": sale.sale_date.date(),
        }
        for sale in customer_sales
    ]

    total_products = Product.objects.filter(user=user).count()
    total_stock_value = Product.objects.filter(user=user).aggregate(total=Sum("stock") * Sum("price"))["total"] or 0

    today_sales = Sale.objects.filter(user=user, created_at__date=date).aggregate(total=Sum("sale_price"))["total"] or 0
    week_sales = (
        Sale.objects.filter(user=user, created_at__date__gte=start_of_week).aggregate(total=Sum("sale_price"))["total"] or 0
    )
    month_sales = (
        Sale.objects.filter(user=user, created_at__date__gte=start_of_month).aggregate(total=Sum("sale_price"))["total"] or 0
    )

    order_status = Order.objects.filter(user=user).values("status").annotate(count=Count("id"))

    return Response(
        {
            "producer_transactions": producer_transactions,
            "customer_transactions": customer_transactions,
            "product_summary": [
                {"title": "Total Products", "value": str(total_products)},
                {"title": "Total Stock Value", "value": f"Rs.{total_stock_value:.2f}"},
            ],
            "sales_overview": [
                {"title": "Total Sales (Today)", "value": f"Rs.{today_sales:.2f}"},
                {"title": "Total Sales (Week)", "value": f"Rs.{week_sales:.2f}"},
                {"title": "Total Sales (Month)", "value": f"Rs.{month_sales:.2f}"},
            ],
            "order_status": [
                {"title": status["status"].capitalize() + " Orders", "value": str(status["count"])}
                for status in order_status
            ],
        }
    )


class DirectSaleViewSet(viewsets.ModelViewSet):
    """
    A viewset for viewing and editing direct sales.
    """

    serializer_class = DirectSaleSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Return direct sales for the current user."""
        return DirectSale.objects.filter(user=self.request.user).select_related("product")

    def perform_create(self, serializer):
        """Set the current user as the seller when creating a direct sale."""
        serializer.save(user=self.request.user)

    @action(detail=False, methods=["get"])
    def recent_sales(self, request):
        """Get recent direct sales (last 10 by default)."""
        limit = int(request.query_params.get("limit", 10))
        sales = self.get_queryset().order_by("-sale_date")[:limit]
        serializer = self.get_serializer(sales, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def sales_summary(self, request):
        """Get summary of direct sales (total sales, count, etc.)."""
        summary = self.get_queryset().aggregate(
            total_sales=Sum("total_amount"),
            total_quantity=Sum("quantity"),
            avg_sale=Sum("total_amount") / Count("id"),
        )

        top_products = (
            self.get_queryset()
            .values("product__name")
            .annotate(total_sold=Sum("quantity"), total_revenue=Sum("total_amount"))
            .order_by("-total_sold")[:5]
        )

        return Response({"summary": summary, "top_products": top_products})


class PurchaseOrderViewSet(viewsets.ModelViewSet):
    serializer_class = PurchaseOrderSerializer

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return PurchaseOrder.objects.none()

        user_profile = getattr(user, "user_profile", None)
        if user_profile:
            return PurchaseOrder.objects.filter(user__user_profile__shop_id=user_profile.shop_id)
        else:
            return PurchaseOrder.objects.none()


# @api_view(["GET"])
# def product_forecasting(request, pk):
#     p = Product.objects.get(pk=pk)
#     actuals, forecast = p.forecast_vs_actual()
#     return Response(
#         {
#             "actuals": actuals,
#             "forecast": forecast,
#             "projected_stockout": p.projected_stockout_date,
#         }
#     )


# @api_view(["GET"])
# def product_seasonality(request, pk):
#     year = int(request.query_params.get("year", timezone.now().year))
#     month = int(request.query_params.get("month", timezone.now().month))
#     start = date(year, month, 1)
#     end = (start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
#     si = Product.objects.get(pk=pk).seasonality_index(start, end)
#     return Response({"seasonality_index": si})


class ShopQRAPIView(APIView):
    def get(self, request, shop_id):
        profile = get_object_or_404(UserProfile, shop_id=shop_id)
        data = ShopQRSerializer(profile, context={"request": request}).data
        return Response(data, status=status.HTTP_200_OK)


class CreatePaymentAPIView(APIView):
    def post(self, request):
        serializer = PaymentSerializer(data=request.data)
        if serializer.is_valid():
            payment = serializer.save(user=request.user, status=Payment.Status.PENDING)
            # fetch shop QR from profile
            profile = request.user.user_profile
            return Response(
                {
                    "payment_id": payment.id,
                    "qr_payload": profile.payment_qr_payload,
                    "qr_url": request.build_absolute_uri(profile.payment_qr_image.url) if profile.payment_qr_image else None,
                },
                status=status.HTTP_201_CREATED,
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PaymentCallbackAPIView(APIView):
    def post(self, request):
        pid = request.data.get("payment_id")
        try:
            payment = Payment.objects.get(pk=pid)
        except Payment.DoesNotExist:
            return Response({"error": "Invalid payment_id"}, status=status.HTTP_400_BAD_REQUEST)

        if request.data.get("status") == "success":
            payment.status = Payment.Status.COMPLETED
            payment.save(update_fields=["status"])
            sale = Sale.objects.create(
                order=payment.order,
                user=payment.user,
                payment=payment,
                quantity=payment.order.quantity,
                sale_price=payment.order.product.listed_price,
            )
            return Response({"sale_id": sale.id}, status=status.HTTP_201_CREATED)
        else:
            payment.status = Payment.Status.FAILED
            payment.save(update_fields=["status"])
            return Response({"detail": "Payment failed"}, status=status.HTTP_200_OK)


class KhaltiInitAPIView(APIView):
    """
    1) Front‑end calls to create a pending Payment and get public key.
    """

    def post(self, request):
        data = KhaltiInitSerializer(data=request.data)
        data.is_valid(raise_exception=True)
        order = get_object_or_404(Order, pk=data.validated_data["order_id"])

        payment = Payment.objects.create(
            order=order, user=request.user, amount=order.total_price, status=Payment.Status.PENDING
        )

        return Response(
            {"payment_id": payment.id, "public_key": settings.KHALTI_PUBLIC_KEY, "amount": float(payment.amount)},
            status=status.HTTP_201_CREATED,
        )


class KhaltiVerifyAPIView(APIView):
    """
    2) Front‑end calls after the Khalti widget returns `token`.
       We hit Khalti’s verify endpoint, validate amount, then record the Sale.
    """

    authentication_classes = []
    permission_classes = []

    def post(self, request):
        data = KhaltiVerifySerializer(data=request.data)
        data.is_valid(raise_exception=True)

        # fetch Payment
        payment = get_object_or_404(Payment, pk=data.validated_data["payment_id"], status=Payment.Status.PENDING)

        # Check amount matches
        if float(payment.amount) != float(data.validated_data["amount"]):
            return Response({"error": "Amount mismatch"}, status=status.HTTP_400_BAD_REQUEST)

        # Verify with Khalti
        resp = requests.post(
            settings.KHALTI_VERIFY_URL,
            data={"token": data.validated_data["token"], "amount": int(data.validated_data["amount"] * 100)},
            headers={"Authorization": f"Key {settings.KHALTI_SECRET_KEY}"},
        )
        if resp.status_code != 200:
            return Response({"error": "Khalti verify failed"}, status=status.HTTP_502_BAD_GATEWAY)

        result = resp.json()
        if result.get("idx"):
            # success!
            payment.status = Payment.Status.COMPLETED
            payment.gateway_token = data.validated_data["token"]
            payment.save(update_fields=["status", "gateway_token"])

            sale = Sale.objects.create(
                order=payment.order,
                user=payment.user,
                payment=payment,
                quantity=payment.order.quantity,
                sale_price=payment.order.product.listed_price,
            )
            return Response({"sale_id": sale.id}, status=status.HTTP_201_CREATED)

        # verification failed on Khalti side
        payment.status = Payment.Status.FAILED
        payment.save(update_fields=["status"])
        return Response({"error": result.get("error_message", "Verification failed")}, status=status.HTTP_400_BAD_REQUEST)


# Category Management ViewSets
class CategoryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing product categories
    """

    queryset = Category.objects.filter(is_active=True).prefetch_related("subcategories", "subcategories__sub_subcategories")
    serializer_class = CategorySerializer

    @action(detail=False, methods=["get"])
    def hierarchy(self, request):
        """Get complete category hierarchy with subcategories and sub-subcategories"""
        categories = Category.objects.filter(is_active=True).prefetch_related("subcategories__sub_subcategories")
        serializer = CategoryHierarchySerializer(categories, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def subcategories(self, request, pk=None):
        """Get subcategories for a specific category (light version - no nested sub-subcategories)"""
        category = self.get_object()
        subcategories = category.subcategories.filter(is_active=True)
        serializer = SubcategoryLightSerializer(subcategories, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def subcategories_full(self, request, pk=None):
        """Get subcategories with full sub-subcategory details"""
        category = self.get_object()
        subcategories = category.subcategories.filter(is_active=True)
        serializer = SubcategorySerializer(subcategories, many=True)
        return Response(serializer.data)


class BrandViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing product brands
    """

    queryset = Brand.objects.filter(is_active=True).prefetch_related("products").order_by("name")
    serializer_class = BrandSerializer
    filterset_class = BrandFilter

    def get_queryset(self):
        """Override to bypass is_active filtering for products action"""
        if self.action == "products":
            # For products action, we need all brands to avoid get_object() failures
            return Brand.objects.prefetch_related("products").order_by("name")
        return super().get_queryset()

    @action(detail=False, methods=["get"])
    def verified(self, request):
        """Get only verified brands"""
        verified_brands = self.get_queryset().filter(is_verified=True)
        serializer = self.get_serializer(verified_brands, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def popular(self, request):
        """Get brands with the most products"""
        from django.db.models import Count, Q

        popular_brands = (
            self.get_queryset()
            .annotate(product_count=Count("products", filter=Q(products__is_active=True)))
            .filter(product_count__gt=0)
            .order_by("-product_count")[:10]
        )

        serializer = self.get_serializer(popular_brands, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def products(self, request, pk=None):
        """Get all products for a specific brand"""
        # Get brand without applying filterset filters
        try:
            brand = Brand.objects.get(pk=pk)
        except Brand.DoesNotExist:
            return Response({"error": f"Brand with ID {pk} not found"}, status=404)

        # Get products for this brand
        products = (
            brand.products.filter(is_active=True)
            .select_related("category", "subcategory", "sub_subcategory", "producer", "user", "location")
            .prefetch_related("images")
        )

        print(f"Total products for brand: {products.count()}")

        # Apply product-level filters from query params
        search = request.query_params.get("search", None)
        print(f"Search parameter: {search}")
        if search:
            products = products.filter(Q(name__icontains=search) | Q(description__icontains=search))
            print(f"Products after search filter: {products.count()}")

        # Apply ordering if specified
        ordering = request.query_params.get("ordering", None)
        if ordering:
            products = products.order_by(ordering)

        paginator = PageNumberPagination()
        paginator.page_size = 20
        page = paginator.paginate_queryset(products, request)

        from .serializers import ProductSerializer

        serializer = ProductSerializer(page, many=True, context={"request": request})
        return paginator.get_paginated_response(serializer.data)


class SubcategoryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing product subcategories
    """

    queryset = Subcategory.objects.filter(is_active=True).select_related("category").prefetch_related("sub_subcategories")

    def get_serializer_class(self):
        """Return light serializer by default, full serializer for detail views"""
        if self.action == "retrieve" or self.request.query_params.get("full", False):
            return SubcategorySerializer
        return SubcategoryLightSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        category_id = self.request.query_params.get("category", None)
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        return queryset

    @action(detail=True, methods=["get"])
    def sub_subcategories(self, request, pk=None):
        """Get sub-subcategories for a specific subcategory"""
        subcategory = self.get_object()
        sub_subcategories = subcategory.sub_subcategories.filter(is_active=True)
        serializer = SubSubcategorySerializer(sub_subcategories, many=True)
        return Response(serializer.data)


class SubSubcategoryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing product sub-subcategories
    """

    queryset = SubSubcategory.objects.filter(is_active=True).select_related("subcategory", "subcategory__category")
    serializer_class = SubSubcategorySerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        subcategory_id = self.request.query_params.get("subcategory", None)
        category_id = self.request.query_params.get("category", None)

        if subcategory_id:
            queryset = queryset.filter(subcategory_id=subcategory_id)
        elif category_id:
            queryset = queryset.filter(subcategory__category_id=category_id)

        return queryset
