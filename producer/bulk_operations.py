import csv
import io
import json
import logging
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

from django.core.cache import cache
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.utils import timezone

logger = logging.getLogger(__name__)

# Try to import optional dependencies
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Excel support disabled.")


@dataclass
class ImportRow:
    """Represents a single row in an import file"""

    row_number: int
    data: Dict[str, Any]
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    is_valid: bool = True
    product_id: Optional[int] = None


@dataclass
class ImportResult:
    """Result of an import operation"""

    job_id: str
    total_rows: int
    success_count: int
    error_count: int
    warning_count: int
    errors: List[Dict]
    created_products: List[int]
    updated_products: List[int]
    failed_rows: List[Dict]
    started_at: datetime
    completed_at: Optional[datetime] = None
    status: str = "processing"  # processing, completed, failed


class ProductImportValidator:
    """Validates product import data"""

    REQUIRED_FIELDS = ["name", "price", "stock"]
    OPTIONAL_FIELDS = [
        "sku",
        "description",
        "cost_price",
        "reorder_level",
        "category_id",
        "subcategory_id",
        "brand_id",
        "size",
        "color",
        "is_active",
    ]

    def __init__(self, user):
        self.user = user
        self.errors = []

    def validate_row(self, row_data: Dict, row_number: int) -> ImportRow:
        """Validate a single row of import data"""
        import_row = ImportRow(row_number=row_number, data=row_data)

        # Check required fields
        for field in self.REQUIRED_FIELDS:
            if field not in row_data or not row_data[field]:
                import_row.errors.append(f"Missing required field: {field}")
                import_row.is_valid = False

        if not import_row.is_valid:
            return import_row

        # Validate name
        name = str(row_data.get("name", "")).strip()
        if len(name) < 2:
            import_row.errors.append("Product name must be at least 2 characters")
            import_row.is_valid = False
        elif len(name) > 255:
            import_row.warnings.append("Product name truncated to 255 characters")
            row_data["name"] = name[:255]

        # Validate price
        try:
            price = self._parse_decimal(row_data.get("price"))
            if price is None or price <= 0:
                import_row.errors.append("Price must be greater than 0")
                import_row.is_valid = False
            elif price > 9999999:
                import_row.warnings.append("Price seems unusually high")
        except (ValueError, InvalidOperation):
            import_row.errors.append(f"Invalid price format: {row_data.get('price')}")
            import_row.is_valid = False

        # Validate stock
        try:
            stock = int(row_data.get("stock", 0))
            if stock < 0:
                import_row.errors.append("Stock cannot be negative")
                import_row.is_valid = False
        except ValueError:
            import_row.errors.append(f"Invalid stock format: {row_data.get('stock')}")
            import_row.is_valid = False

        # Validate cost_price if provided
        if row_data.get("cost_price"):
            try:
                cost_price = self._parse_decimal(row_data.get("cost_price"))
                if cost_price and cost_price < 0:
                    import_row.errors.append("Cost price cannot be negative")
                    import_row.is_valid = False
                elif cost_price and price and cost_price > price:
                    import_row.warnings.append("Cost price is higher than selling price")
            except (ValueError, InvalidOperation):
                import_row.errors.append(f"Invalid cost_price format: {row_data.get('cost_price')}")
                import_row.is_valid = False

        # Validate category if provided
        if row_data.get("category_id"):
            from .models import Category

            try:
                category_id = int(row_data["category_id"])
                if not Category.objects.filter(id=category_id).exists():
                    import_row.errors.append(f"Category ID {category_id} does not exist")
                    import_row.is_valid = False
            except ValueError:
                import_row.errors.append(f"Invalid category_id: {row_data['category_id']}")
                import_row.is_valid = False

        # Validate brand if provided
        if row_data.get("brand_id"):
            from .models import Brand

            try:
                brand_id = int(row_data["brand_id"])
                if not Brand.objects.filter(id=brand_id).exists():
                    import_row.errors.append(f"Brand ID {brand_id} does not exist")
                    import_row.is_valid = False
            except ValueError:
                import_row.errors.append(f"Invalid brand_id: {row_data['brand_id']}")
                import_row.is_valid = False

        return import_row

    def _parse_decimal(self, value) -> Optional[Decimal]:
        """Parse a decimal value from various formats"""
        if value is None or value == "":
            return None

        if isinstance(value, (int, float)):
            return Decimal(str(value))

        # Clean string value
        value_str = str(value).replace(",", "").replace("$", "").replace("₹", "").strip()
        return Decimal(value_str)


class ProductImporter:
    """Handles bulk product imports"""

    def __init__(self, user, update_existing: bool = True):
        self.user = user
        self.update_existing = update_existing
        self.validator = ProductImportValidator(user)

    def import_csv(self, file_content: str, job_id: str) -> ImportResult:
        """Import products from CSV content"""
        result = ImportResult(
            job_id=job_id,
            total_rows=0,
            success_count=0,
            error_count=0,
            warning_count=0,
            errors=[],
            created_products=[],
            updated_products=[],
            failed_rows=[],
            started_at=timezone.now(),
        )

        try:
            # Parse CSV
            reader = csv.DictReader(io.StringIO(file_content))
            rows = list(reader)
            result.total_rows = len(rows)

            # Process in batches
            batch_size = 100
            for i in range(0, len(rows), batch_size):
                batch = rows[i : i + batch_size]
                self._process_batch(batch, i + 1, result)

                # Update progress in cache
                progress = {
                    "job_id": job_id,
                    "status": "processing",
                    "processed": min(i + batch_size, len(rows)),
                    "total": len(rows),
                    "percent": int((min(i + batch_size, len(rows)) / len(rows)) * 100),
                }
                cache.set(f"import_progress_{job_id}", progress, 3600)

            result.status = "completed"
            result.completed_at = timezone.now()

        except Exception as e:
            logger.exception(f"Import failed for job {job_id}")
            result.status = "failed"
            result.errors.append({"error": str(e)})

        return result

    def import_excel(self, file_content: bytes, job_id: str) -> ImportResult:
        """Import products from Excel file"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel import")

        result = ImportResult(
            job_id=job_id,
            total_rows=0,
            success_count=0,
            error_count=0,
            warning_count=0,
            errors=[],
            created_products=[],
            updated_products=[],
            failed_rows=[],
            started_at=timezone.now(),
        )

        try:
            # Load workbook
            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in ws[1]]

            # Convert to list of dicts
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        row_dict[header] = row[i]
                rows.append(row_dict)

            result.total_rows = len(rows)

            # Process in batches
            batch_size = 100
            for i in range(0, len(rows), batch_size):
                batch = rows[i : i + batch_size]
                self._process_batch(batch, i + 1, result)

                # Update progress
                progress = {
                    "job_id": job_id,
                    "status": "processing",
                    "processed": min(i + batch_size, len(rows)),
                    "total": len(rows),
                    "percent": int((min(i + batch_size, len(rows)) / len(rows)) * 100),
                }
                cache.set(f"import_progress_{job_id}", progress, 3600)

            result.status = "completed"
            result.completed_at = timezone.now()

        except Exception as e:
            logger.exception(f"Excel import failed for job {job_id}")
            result.status = "failed"
            result.errors.append({"error": str(e)})

        return result

    def _process_batch(self, rows: List[Dict], start_row: int, result: ImportResult):
        """Process a batch of rows"""
        from .models import Brand, Category, Product

        import_rows = []

        # Validate all rows first
        for i, row_data in enumerate(rows):
            row_number = start_row + i
            import_row = self.validator.validate_row(row_data, row_number)
            import_rows.append(import_row)

            if import_row.warnings:
                result.warning_count += len(import_row.warnings)

        # Process valid rows
        valid_rows = [r for r in import_rows if r.is_valid]

        with transaction.atomic():
            for import_row in valid_rows:
                try:
                    product = self._create_or_update_product(import_row.data)

                    if import_row.product_id:
                        result.updated_products.append(product.id)
                    else:
                        result.created_products.append(product.id)

                    result.success_count += 1

                except Exception as e:
                    logger.exception(f"Error processing row {import_row.row_number}")
                    import_row.errors.append(str(e))
                    import_row.is_valid = False
                    result.error_count += 1
                    result.failed_rows.append(
                        {"row": import_row.row_number, "data": import_row.data, "errors": import_row.errors}
                    )

        # Track invalid rows
        invalid_rows = [r for r in import_rows if not r.is_valid]
        for import_row in invalid_rows:
            result.error_count += len(import_row.errors)
            result.failed_rows.append({"row": import_row.row_number, "data": import_row.data, "errors": import_row.errors})

    def _create_or_update_product(self, data: Dict) -> "Product":
        """Create or update a product from import data"""
        from .models import Product

        # Check if product exists by SKU
        sku = data.get("sku")
        existing_product = None

        if sku:
            existing_product = Product.objects.filter(sku=sku, user=self.user).first()

        # Also check by name
        if not existing_product and self.update_existing:
            existing_product = Product.objects.filter(name__iexact=data["name"], user=self.user).first()

        # Prepare product data
        product_data = {
            "name": data["name"].strip(),
            "price": self.validator._parse_decimal(data.get("price")),
            "stock": int(data.get("stock", 0)),
            "user": self.user,
        }

        # Optional fields
        if data.get("sku"):
            product_data["sku"] = str(data["sku"]).strip()
        if data.get("description"):
            product_data["description"] = str(data["description"])
        if data.get("cost_price"):
            product_data["cost_price"] = float(self.validator._parse_decimal(data["cost_price"]))
        if data.get("reorder_level"):
            product_data["reorder_level"] = int(data["reorder_level"])
        if data.get("category_id"):
            product_data["category_id"] = int(data["category_id"])
        if data.get("subcategory_id"):
            product_data["subcategory_id"] = int(data["subcategory_id"])
        if data.get("brand_id"):
            product_data["brand_id"] = int(data["brand_id"])
        if data.get("size"):
            product_data["size"] = data["size"]
        if data.get("color"):
            product_data["color"] = data["color"]
        if "is_active" in data:
            product_data["is_active"] = str(data["is_active"]).lower() in ["true", "1", "yes", "active"]

        if existing_product:
            # Update existing
            for key, value in product_data.items():
                setattr(existing_product, key, value)
            existing_product.save()
            return existing_product
        else:
            # Create new
            return Product.objects.create(**product_data)


class ProductExporter:
    """Handles bulk product exports"""

    EXPORT_FIELDS = [
        "id",
        "sku",
        "name",
        "description",
        "price",
        "cost_price",
        "stock",
        "reorder_level",
        "category__name",
        "subcategory__name",
        "brand__name",
        "size",
        "color",
        "is_active",
        "created_at",
    ]

    def __init__(self, user):
        self.user = user

    def export_csv(self, filters: Optional[Dict] = None) -> Tuple[str, str]:
        """
        Export products to CSV.
        Returns (filename, content)
        """
        products = self._get_products(filters)

        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        headers = [h.replace("__", "_") for h in self.EXPORT_FIELDS]
        writer.writerow(headers)

        # Write data
        for product in products:
            row = self._product_to_row(product)
            writer.writerow(row)

        filename = f"products_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return filename, output.getvalue()

    def export_excel(self, filters: Optional[Dict] = None) -> Tuple[str, bytes]:
        """
        Export products to Excel.
        Returns (filename, content)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        products = self._get_products(filters)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        headers = [h.replace("__", "_").title() for h in self.EXPORT_FIELDS]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_num, product in enumerate(products, 2):
            row = self._product_to_row(product)
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Alternate row colors
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in range(1, len(products) + 2):
                cell = ws[f"{column}{row}"]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"products_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return filename, output.getvalue()

    def _get_products(self, filters: Optional[Dict]):
        """Get products queryset with optional filters"""
        from .models import Product

        qs = Product.objects.filter(user=self.user).select_related("category", "subcategory", "brand")

        if filters:
            if filters.get("category_id"):
                qs = qs.filter(category_id=filters["category_id"])
            if filters.get("is_active") is not None:
                qs = qs.filter(is_active=filters["is_active"])
            if filters.get("stock_status") == "in_stock":
                qs = qs.filter(stock__gt=0)
            elif filters.get("stock_status") == "out_of_stock":
                qs = qs.filter(stock=0)

        return qs

    def _product_to_row(self, product) -> List:
        """Convert product to list of values"""
        return [
            product.id,
            product.sku or "",
            product.name,
            product.description or "",
            product.price,
            product.cost_price or "",
            product.stock,
            product.reorder_level,
            product.category.name if product.category else "",
            product.subcategory.name if product.subcategory else "",
            product.brand.name if product.brand else "",
            product.size or "",
            product.color or "",
            "Yes" if product.is_active else "No",
            product.created_at.strftime("%Y-%m-%d %H:%M") if product.created_at else "",
        ]

    def generate_template(self, format: str = "csv") -> Tuple[str, str]:
        """Generate import template file"""
        if format == "csv":
            return self._generate_csv_template()
        elif format == "excel":
            return self._generate_excel_template()
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _generate_csv_template(self) -> Tuple[str, str]:
        """Generate CSV import template with sample data"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Headers
        headers = [
            "name",
            "sku",
            "description",
            "price",
            "cost_price",
            "stock",
            "reorder_level",
            "category_id",
            "subcategory_id",
            "brand_id",
            "size",
            "color",
            "is_active",
        ]
        writer.writerow(headers)

        # Sample data row
        sample = [
            "Sample Product Name",
            "SKU-001",
            "Product description here",
            "999.99",
            "599.99",
            "100",
            "10",
            "1",
            "",
            "",
            "M",
            "Blue",
            "Yes",
        ]
        writer.writerow(sample)

        return "product_import_template.csv", output.getvalue()

    def _generate_excel_template(self) -> Tuple[str, bytes]:
        """Generate Excel import template"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Template"

        # Headers
        headers = [
            "name",
            "sku",
            "description",
            "price",
            "cost_price",
            "stock",
            "reorder_level",
            "category_id",
            "subcategory_id",
            "brand_id",
            "size",
            "color",
            "is_active",
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        required_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

            # Highlight required fields
            if header in ["name", "price", "stock"]:
                cell.fill = required_fill

        # Add instructions sheet
        instructions = wb.create_sheet("Instructions")
        instructions.column_dimensions["A"].width = 30
        instructions.column_dimensions["B"].width = 60

        instruction_data = [
            ("Field", "Description"),
            ("name", "Product name (required)"),
            ("sku", "Unique product code (optional)"),
            ("description", "Product description (optional)"),
            ("price", "Selling price (required, numeric)"),
            ("cost_price", "Cost price (optional, numeric)"),
            ("stock", "Current stock quantity (required, numeric)"),
            ("reorder_level", "Stock level to trigger reorder (optional)"),
            ("category_id", "Category ID number (optional)"),
            ("subcategory_id", "Subcategory ID number (optional)"),
            ("brand_id", "Brand ID number (optional)"),
            ("size", "Product size (optional)"),
            ("color", "Product color (optional)"),
            ("is_active", "Yes/No (optional, default: Yes)"),
            ("", ""),
            ("Notes:", ""),
            ("- Yellow highlighted fields are required", ""),
            ("- Existing products will be updated if SKU matches", ""),
            ("- New products will be created for new SKUs", ""),
        ]

        for row_num, (field, desc) in enumerate(instruction_data, 1):
            instructions.cell(row=row_num, column=1, value=field)
            instructions.cell(row=row_num, column=2, value=desc)

        # Save
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return "product_import_template.xlsx", output.read()


def get_import_progress(job_id: str) -> Optional[Dict]:
    """Get the progress of an import job"""
    return cache.get(f"import_progress_{job_id}")


def get_import_result(job_id: str) -> Optional[ImportResult]:
    """Get the result of a completed import"""
    result_data = cache.get(f"import_result_{job_id}")
    if result_data:
        # Convert dict back to ImportResult
        return ImportResult(**result_data)
    return None
