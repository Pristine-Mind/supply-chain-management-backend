import uuid

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .bulk_operations import ProductExporter, get_import_progress, get_import_result
from .tasks_bulk import (
    export_products_task,
    import_products_csv_task,
    import_products_excel_task,
)


class BulkImportView(APIView):
    """
    Upload and import products from CSV or Excel file.
    Processing happens asynchronously via Celery.

    Supports:
    - CSV files (.csv)
    - Excel files (.xlsx, .xls)
    """

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        """Upload file and start import"""
        if "file" not in request.FILES:
            return Response(
                {"error": 'No file provided. Use multipart/form-data with "file" field.'}, status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES["file"]
        update_existing = request.data.get("update_existing", "true").lower() == "true"

        # Validate file type
        filename = file.name.lower()
        if not (filename.endswith(".csv") or filename.endswith(".xlsx") or filename.endswith(".xls")):
            return Response(
                {"error": "Invalid file type. Only CSV and Excel files are supported."}, status=status.HTTP_400_BAD_REQUEST
            )

        # Generate job ID
        job_id = str(uuid.uuid4())

        try:
            if filename.endswith(".csv"):
                # Read CSV content
                content = file.read().decode("utf-8")

                # Start async task
                task = import_products_csv_task.delay(
                    file_content=content, user_id=request.user.id, job_id=job_id, update_existing=update_existing
                )
            else:
                # Read Excel content
                content = file.read()

                # Start async task
                task = import_products_excel_task.delay(
                    file_content_bytes=content, user_id=request.user.id, job_id=job_id, update_existing=update_existing
                )

            return Response(
                {
                    "success": True,
                    "message": "Import started successfully",
                    "job_id": job_id,
                    "task_id": task.id,
                    "status": "processing",
                    "check_status_url": f"/api/v1/producer/import/{job_id}/status/",
                },
                status=status.HTTP_202_ACCEPTED,
            )

        except Exception as e:
            return Response({"error": f"Failed to start import: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ImportStatusView(APIView):
    """
    Check the status of an import job.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, job_id):
        # Get progress
        progress = get_import_progress(job_id)

        if not progress:
            # Check if completed
            result = get_import_result(job_id)
            if result:
                return Response(
                    {
                        "job_id": job_id,
                        "status": "completed",
                        "progress": {"processed": result.total_rows, "total": result.total_rows, "percent": 100},
                        "result": {
                            "total_rows": result.total_rows,
                            "success_count": result.success_count,
                            "error_count": result.error_count,
                            "warning_count": result.warning_count,
                            "created_products": len(result.created_products),
                            "updated_products": len(result.updated_products),
                        },
                    }
                )

            return Response({"error": "Job not found"}, status=status.HTTP_404_NOT_FOUND)

        return Response(
            {
                "job_id": job_id,
                "status": progress.get("status"),
                "progress": {
                    "processed": progress.get("processed", 0),
                    "total": progress.get("total", 0),
                    "percent": progress.get("percent", 0),
                },
                "counts": (
                    {"success": progress.get("success_count"), "errors": progress.get("error_count")}
                    if "success_count" in progress
                    else None
                ),
            }
        )


class ImportResultView(APIView):
    """
    Get detailed results of a completed import job.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, job_id):
        result_data = get_import_result(job_id)

        if not result_data:
            # Check if still processing
            progress = get_import_progress(job_id)
            if progress and progress.get("status") == "processing":
                return Response({"job_id": job_id, "status": "processing", "message": "Import is still in progress"})

            return Response({"error": "Job not found or expired"}, status=status.HTTP_404_NOT_FOUND)

        # Return full result
        response_data = {
            "job_id": result_data.get("job_id"),
            "status": result_data.get("status"),
            "summary": {
                "total_rows": result_data.get("total_rows"),
                "success_count": result_data.get("success_count"),
                "error_count": result_data.get("error_count"),
                "warning_count": result_data.get("warning_count"),
                "created_products": len(result_data.get("created_products", [])),
                "updated_products": len(result_data.get("updated_products", [])),
            },
            "timing": {"started_at": result_data.get("started_at"), "completed_at": result_data.get("completed_at")},
        }

        # Include errors if requested
        if request.query_params.get("include_errors", "").lower() == "true":
            failed_rows = result_data.get("failed_rows", [])
            response_data["errors"] = {
                "total_errors": len(failed_rows),
                "failed_rows": failed_rows[:50],  # Limit to first 50
            }

        return Response(response_data)


class BulkExportView(APIView):
    """
    Export products to CSV or Excel.
    Processing happens asynchronously via Celery.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Start export job"""
        format_type = request.data.get("format", "csv").lower()
        filters = request.data.get("filters", {})

        if format_type not in ["csv", "excel"]:
            return Response({"error": 'Invalid format. Use "csv" or "excel".'}, status=status.HTTP_400_BAD_REQUEST)

        # Generate job ID
        job_id = str(uuid.uuid4())

        try:
            # Start async task
            task = export_products_task.delay(user_id=request.user.id, filters=filters, format=format_type, job_id=job_id)

            return Response(
                {
                    "success": True,
                    "message": "Export started successfully",
                    "job_id": job_id,
                    "task_id": task.id,
                    "status": "processing",
                    "check_status_url": f"/api/v1/producer/export/{job_id}/status/",
                },
                status=status.HTTP_202_ACCEPTED,
            )

        except Exception as e:
            return Response({"error": f"Failed to start export: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get(self, request):
        """Quick export (synchronous for small datasets)"""
        format_type = request.query_params.get("format", "csv").lower()

        filters = {}
        if "category_id" in request.query_params:
            filters["category_id"] = int(request.query_params["category_id"])
        if "is_active" in request.query_params:
            filters["is_active"] = request.query_params["is_active"].lower() == "true"

        try:
            exporter = ProductExporter(request.user)

            if format_type == "csv":
                filename, content = exporter.export_csv(filters)
                response = Response(content, content_type="text/csv")
                response["Content-Disposition"] = f'attachment; filename="{filename}"'
                return response
            elif format_type == "excel":
                filename, content = exporter.export_excel(filters)
                response = Response(
                    content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = f'attachment; filename="{filename}"'
                return response
            else:
                return Response({"error": 'Invalid format. Use "csv" or "excel".'}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"error": f"Export failed: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExportStatusView(APIView):
    """
    Check status of an export job and get download URL.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, job_id):
        from django.core.cache import cache

        progress = cache.get(f"export_progress_{job_id}")
        result = cache.get(f"export_result_{job_id}")

        if not progress and not result:
            return Response({"error": "Job not found or expired"}, status=status.HTTP_404_NOT_FOUND)

        if result:
            return Response(
                {
                    "job_id": job_id,
                    "status": "completed",
                    "download": {
                        "url": result.get("download_url"),
                        "filename": result.get("filename"),
                        "format": result.get("format"),
                        "size_bytes": result.get("file_size"),
                    },
                }
            )

        return Response(
            {"job_id": job_id, "status": progress.get("status"), "progress": {"percent": progress.get("percent", 0)}}
        )


class ImportTemplateView(APIView):
    """
    Download import template file.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        format_type = request.query_params.get("format", "csv").lower()

        try:
            exporter = ProductExporter(request.user)
            filename, content = exporter.generate_template(format_type)

            if format_type == "csv":
                response = Response(content, content_type="text/csv")
            else:
                response = Response(
                    content, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            return Response(
                {"error": f"Failed to generate template: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ImportValidationView(APIView):
    """
    Validate import file without actually importing.
    Returns validation errors and warnings.
    """

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if "file" not in request.FILES:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES["file"]

        try:
            from .bulk_operations import ProductImporter

            importer = ProductImporter(request.user, update_existing=False)

            # Read and validate first 100 rows
            filename = file.name.lower()

            if filename.endswith(".csv"):
                content = file.read().decode("utf-8")
                import csv
                import io

                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)[:100]
            elif filename.endswith((".xlsx", ".xls")):
                try:
                    import openpyxl

                    wb = openpyxl.load_workbook(io.BytesIO(file.read()))
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    rows = []
                    for row in ws.iter_rows(min_row=2, max_row=101, values_only=True):
                        rows.append({headers[i]: row[i] for i in range(len(headers))})
                except ImportError:
                    return Response(
                        {"error": "Excel validation not available"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
            else:
                return Response({"error": "Invalid file type"}, status=status.HTTP_400_BAD_REQUEST)

            # Validate each row
            errors = []
            warnings = []
            valid_count = 0

            for i, row in enumerate(rows, 1):
                import_row = importer.validator.validate_row(row, i)

                if import_row.errors:
                    errors.append({"row": i, "errors": import_row.errors})
                elif import_row.warnings:
                    warnings.append({"row": i, "warnings": import_row.warnings})
                    valid_count += 1
                else:
                    valid_count += 1

            return Response(
                {
                    "valid": len(errors) == 0,
                    "summary": {
                        "total_rows_checked": len(rows),
                        "valid_rows": valid_count,
                        "rows_with_errors": len(errors),
                        "rows_with_warnings": len(warnings),
                    },
                    "errors": errors[:20],  # Limit errors shown
                    "warnings": warnings[:20],
                    "sample_valid_row": rows[0] if rows and len(errors) == 0 else None,
                }
            )

        except Exception as e:
            return Response({"error": f"Validation failed: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def import_export_stats(request):
    """
    Get statistics about past import/export operations.
    """
    from django.core.cache import cache

    # This is a simple implementation - could be enhanced with database tracking
    return Response(
        {
            "note": "Import/export operations are tracked per-job and expire after 24 hours",
            "endpoints": {
                "import": "/api/v1/producer/import/",
                "export": "/api/v1/producer/export/",
                "template": "/api/v1/producer/import/template/",
                "validate": "/api/v1/producer/import/validate/",
            },
        }
    )
