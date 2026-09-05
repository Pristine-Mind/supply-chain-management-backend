import gc
from asyncio.log import logger
from datetime import date, datetime

import torch
from django.utils.timezone import is_naive, localtime, make_aware
from keybert import KeyBERT
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from producer.models import MarketplaceProduct

# Lazy-load KeyBERT model to avoid memory bloat at startup
_kw_model = None


def get_keybert_model():
    """Lazy-load KeyBERT model on first use to avoid memory issues."""
    global _kw_model
    if _kw_model is None:
        logger.info("Loading KeyBERT model...")
        _kw_model = KeyBERT()
    return _kw_model


def unload_keybert_model():
    """Unload the KeyBERT model to free memory."""
    global _kw_model
    if _kw_model is not None:
        logger.info("Unloading KeyBERT model to free memory...")
        try:
            # Try to move model to CPU and clear CUDA cache if available
            if hasattr(_kw_model, "model") and hasattr(_kw_model.model, "cpu"):
                _kw_model.model.cpu()
            if torch.cuda.is_available():
                torch.cuda.empty_cache()
        except:
            pass
        _kw_model = None
        gc.collect()


def export_queryset_to_excel(queryset, field_names, headers=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Export"

    if headers:
        ws.append(headers)
    else:
        ws.append([field.replace("_", " ").title() for field in field_names])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for obj in queryset:
        row = []
        for field in field_names:
            value = getattr(obj, field)
            if callable(value):
                value = value()
            elif hasattr(value, "all"):
                value = ", ".join([str(item) for item in value.all()])
            elif isinstance(value, datetime):
                # Handle timezone for datetime objects
                if is_naive(value):
                    value = make_aware(value)
                value = localtime(value).strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(value, date):
                # Handle date objects (without time)
                value = value.strftime("%Y-%m-%d")
            else:
                # Convert model instances and other objects to strings
                value = str(value) if value is not None else ""
            row.append(value)
        ws.append(row)

    for col_num, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_num)].width = length + 5

    return wb


def generate_and_save_product_tags(product_id):
    """
    Extracts semantic search tags using KeyBERT and saves them back to the product.
    Runs asynchronously to prevent blocking the web server.
    Overwrites any existing tags with the new optimized format.
    """
    try:
        product = MarketplaceProduct.objects.select_related("product", "product__brand", "product__category").get(
            id=product_id
        )
        base_product = product.product

        name = base_product.name if base_product and base_product.name else ""
        desc = base_product.description if base_product and base_product.description else ""
        text_to_analyze = f"{name} {desc}".strip()

        if not text_to_analyze:
            logger.warning(f"No text to analyze for product {product_id}")
            return []

        static_tags = []
        if base_product and getattr(base_product, "brand", None):
            b_name = getattr(base_product.brand, "name", None) or str(base_product.brand)
            b_clean = b_name.strip().lower() if b_name else ""
            if b_clean and not b_clean.startswith("dummy_") and b_clean != "unbranded":
                static_tags.append(b_clean)

        if not static_tags:
            try:
                from .models import Brand

                for b_name in Brand.objects.values_list("name", flat=True):
                    b_clean = b_name.strip().lower() if b_name else ""
                    if b_clean and len(b_clean) > 1 and b_clean != "unbranded" and b_clean in text_to_analyze.lower():
                        static_tags.append(b_clean)
                        break
            except Exception:
                pass

        if not static_tags and name:
            first_word = name.split()[0].strip().lower()
            if first_word.isalpha() and len(first_word) >= 2 and first_word != "unbranded":
                static_tags.append(first_word)
        if base_product and getattr(base_product, "category", None):
            c_name = getattr(base_product.category, "name", None)
            if c_name:
                static_tags.append(str(c_name).strip().lower())
        kw_model = get_keybert_model()
        keywords = kw_model.extract_keywords(text_to_analyze, keyphrase_ngram_range=(1, 2), stop_words="english", top_n=5)

        nlp_tags = [kw[0] for kw in keywords]
        combined_tags = []
        for tag in static_tags + nlp_tags:
            clean_tag = tag.strip().lower()
            if clean_tag and clean_tag not in combined_tags:
                combined_tags.append(clean_tag)
        product.search_tags = combined_tags
        product.save(update_fields=["search_tags"])

        logger.info(f"AI tagged product {product_id} with: {combined_tags}")
        return combined_tags

    except MarketplaceProduct.DoesNotExist:
        logger.error(f"MarketplaceProduct {product_id} not found.")
        return None
    except Exception as e:
        logger.error(f"AI Extraction failed for product {product_id}: {str(e)}")
        return None
