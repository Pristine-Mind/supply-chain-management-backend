import datetime
import logging
from io import BytesIO

from django.contrib import admin, messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils.exceptions import SheetTitleException

from report.models import DailySalesReport, DailySalesReportItem
from user.admin_mixins import RoleBasedAdminMixin

logger = logging.getLogger(__name__)


def safe_sheet_title(title: str) -> str:
    for ch in ["\n", "\r", "\t", "/", "\\", "?", "*", "[", "]", ":"]:
        title = title.replace(ch, " ")
    return title.strip()[:31] or "Sheet1"


def export_to_excel(queryset, filename_prefix, headers, row_iterator):
    """
    Streaming export via write-only mode—no ws[0] access.
    """
    wb = Workbook(write_only=True)
    try:
        ws = wb.create_sheet(title=safe_sheet_title(filename_prefix))
    except SheetTitleException:
        ws = wb.create_sheet(title="Sheet1")

    ws.append(headers)

    for count, row in enumerate(row_iterator(), start=1):
        ws.append(row)
        if count % 1000 == 0:
            logger.info(f"Exported {count} rows for {filename_prefix}")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.date.today().strftime("%Y%m%d")
    fname = f"{filename_prefix}_{today}.xlsx"
    resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@admin.register(DailySalesReport)
class DailySalesReportAdmin(RoleBasedAdminMixin, admin.ModelAdmin):
    required_role = "admin"  # Only admin users can access
    list_display = ("report_date", "generated_at", "total_items", "total_revenue")
    actions = ["export_reports"]

    def export_reports(self, request, queryset):
        if not queryset:
            self.message_user(request, "No reports selected.", level=messages.WARNING)
            return
        headers = ["Report Date", "Generated At", "Total Items", "Total Revenue"]

        def rows():
            for rpt in queryset.order_by("report_date"):
                yield [
                    rpt.report_date.strftime("%Y-%m-%d"),
                    rpt.generated_at.strftime("%Y-%m-%d %H:%M:%S"),
                    rpt.total_items,
                    float(rpt.total_revenue),
                ]

        try:
            return export_to_excel(queryset, "daily_sales_reports", headers, rows)
        except Exception as e:
            logger.exception("Failed to export reports")
            self.message_user(request, f"Error: {e}", level=messages.ERROR)

    export_reports.short_description = "Export selected reports to Excel"


@admin.register(DailySalesReportItem)
class DailySalesReportItemAdmin(RoleBasedAdminMixin, admin.ModelAdmin):
    required_role = "admin"  # Only admin users can access
    list_display = ("report", "date", "product", "product_owner", "customer", "unit_price", "quantity", "line_total")
    list_select_related = ("report", "product", "product_owner", "customer")
    actions = ["export_items"]

    def export_items(self, request, queryset):
        if not queryset:
            self.message_user(request, "No items selected.", level=messages.WARNING)
            return
        headers = ["Report Date", "Date", "Product", "Product Owner", "Customer", "Unit Price", "Quantity", "Line Total"]

        def rows():
            for item in queryset.select_related("report", "product", "product_owner", "customer"):
                yield [
                    item.report.report_date.strftime("%Y-%m-%d"),
                    item.date.strftime("%Y-%m-%d"),
                    str(item.product.product.name),
                    item.product_owner.username if item.product_owner else "-",
                    item.customer.username if item.customer else "Guest",
                    float(item.unit_price),
                    item.quantity,
                    float(item.line_total),
                ]

        try:
            return export_to_excel(queryset, "daily_sales_items", headers, rows)
        except Exception as e:
            logger.exception("Failed to export items")
            self.message_user(request, f"Error: {e}", level=messages.ERROR)

    export_items.short_description = "Export selected items to Excel"
